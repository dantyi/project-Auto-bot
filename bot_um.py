import subprocess
import psutil
import threading
import time
import pyautogui
import os
import openpyxl
import pyperclip
import win32clipboard

# ==========================================
# CONFIG
# ==========================================
CRM_PATH = r"C:\Users\dell\AppData\Roaming\Microsoft\Windows\Start Menu\Programs\Global Hitss\Global Hitss\Nuevo CRM.appref-ms"
IMAGEN_CONSULTAS        = r"C:\Users\dell\Desktop\kickoff\project-Auto-bot\stratic\consultas.png"
IMAGEN_FLUJO_TRABAJO    = r"C:\Users\dell\Desktop\kickoff\project-Auto-bot\stratic\flujo de trabajo.png"
IMAGEN_CRM              = r"C:\Users\dell\Desktop\kickoff\project-Auto-bot\stratic\CRM.png"
IMAGEN_OTP_OCUPADA      = r"C:\Users\dell\Desktop\kickoff\project-Auto-bot\stratic\OTPOCUPADA.png"
IMAGEN_TAREAS           = r"C:\Users\dell\Desktop\kickoff\project-Auto-bot\stratic\tareas.png"
IMAGEN_ANOTACIONES      = r"C:\Users\dell\Desktop\kickoff\project-Auto-bot\stratic\anotaciones.png"
IMAGEN_VISITA_TECNICA   = r"C:\Users\dell\Desktop\kickoff\project-Auto-bot\stratic\Visita_tecnica.png"
IMAGEN_OTH_INSTALACION  = r"C:\Users\dell\Desktop\kickoff\project-Auto-bot\stratic\oth_instalacion.png"
IMAGEN_OTH_DISEÑO_UM    = r"C:\Users\dell\Desktop\kickoff\project-Auto-bot\stratic\oth_diseño_um.png"
IMAGEN_OTH_ACOMETIDA    = r"C:\Users\dell\Desktop\kickoff\project-Auto-bot\stratic\oth_acometida.png"
IMAGEN_FACTURACION        = r"C:\Users\dell\Desktop\kickoff\project-Auto-bot\stratic\facturacion.png"
IMAGEN_EDITAR_INCIDENTE   = r"C:\Users\dell\Desktop\kickoff\project-Auto-bot\stratic\editar_incidente.png"
IMAGEN_GUARDAR            = r"C:\Users\dell\Desktop\kickoff\project-Auto-bot\stratic\guardar.png"
IMAGEN_GUARDAR_OTP        = r"C:\Users\dell\Desktop\kickoff\project-Auto-bot\stratic\guardar_otp.png"
IMAGEN_CIERRE             = r"C:\Users\dell\Desktop\kickoff\project-Auto-bot\stratic\cierre.png"
IMAGEN_EDITAR_TAREA       = r"C:\Users\dell\Desktop\kickoff\project-Auto-bot\stratic\editar_tarea.png"
IMAGEN_ASIGNACION_TAREA   = r"C:\Users\dell\Desktop\kickoff\project-Auto-bot\stratic\asignacion_tarea.png"
IMAGEN_CODIGOS_RESOLUCION = r"C:\Users\dell\Desktop\kickoff\project-Auto-bot\stratic\codigos_de_resolucion.png"
IMAGEN_COD_RESOLUCION_1   = r"C:\Users\dell\Desktop\kickoff\project-Auto-bot\stratic\codigo_de_resolucion_1.png"

EXCEL_PATH    = r"C:\Users\dell\Desktop\kickoff\project-Auto-bot\datos.xlsx"
ESTADO_COLUMNA = "COMPLETADO"

INACTIVITY_LIMIT = 480
CONFIDENCE = 0.87

crm_process = None
last_activity_time = time.time()
flujo_en_ejecucion = False
lock = threading.Lock()

fila_actual_excel = None

pyautogui.FAILSAFE = True

# ==========================================
# LIMITE CRM (900 chars, Enter = 4)
# ==========================================
MAX_CARACTERES_CRM = 1000
ENTER_COST = 4

def dividir_texto_por_limite(texto, limite=MAX_CARACTERES_CRM):
    if texto is None:
        return []
    texto = str(texto).replace("\r\n", "\n").replace("\r", "\n")
    partes = []
    actual = []
    costo_actual = 0

    for linea in texto.split("\n"):
        c_linea = len(linea)
        c_enter = ENTER_COST if len(actual) > 0 else 0

        if (costo_actual + c_enter + c_linea) > limite and len(actual) > 0:
            partes.append("\n".join(actual))
            actual = []
            costo_actual = 0
            c_enter = 0

        if c_linea > limite:
            if len(actual) > 0:
                partes.append("\n".join(actual))
                actual = []
                costo_actual = 0
            start = 0
            while start < len(linea):
                end = min(start + limite, len(linea))
                partes.append(linea[start:end])
                start = end
            continue

        if c_enter:
            costo_actual += c_enter
        actual.append(linea)
        costo_actual += c_linea

    if len(actual) > 0:
        partes.append("\n".join(actual))

    return partes

# ====== LOCK PROPIO (evita 2 instancias de bot_um.py) ======
LOCK_FILE = os.path.join(os.path.dirname(__file__), "bot_um.lock")
lock_fd = None

def adquirir_lock():
    global lock_fd
    try:
        lock_fd = os.open(LOCK_FILE, os.O_CREAT | os.O_EXCL | os.O_RDWR)
        os.write(lock_fd, str(os.getpid()).encode())
        return True
    except FileExistsError:
        return False

def liberar_lock():
    global lock_fd
    try:
        if lock_fd is not None:
            os.close(lock_fd)
            lock_fd = None
        if os.path.exists(LOCK_FILE):
            os.remove(LOCK_FILE)
    except:
        pass

# ====== LOCK DE COLA COMPARTIDA (bot.py + bot_um.py) ======
COLA_LOCK_FILE = os.path.join(os.path.dirname(__file__), "cola.lock")
cola_lock_fd = None

def adquirir_cola_lock():
    global cola_lock_fd
    intentos = 0
    while True:
        try:
            cola_lock_fd = os.open(COLA_LOCK_FILE, os.O_CREAT | os.O_EXCL | os.O_RDWR)
            os.write(cola_lock_fd, str(os.getpid()).encode())
            return True
        except FileExistsError:
            if intentos % 10 == 0:
                print("⏳ Cola ocupada por otro bot, esperando...")
            time.sleep(1)
            intentos += 1

def liberar_cola_lock():
    global cola_lock_fd
    try:
        if cola_lock_fd is not None:
            os.close(cola_lock_fd)
            cola_lock_fd = None
        if os.path.exists(COLA_LOCK_FILE):
            os.remove(COLA_LOCK_FILE)
    except:
        pass

# ==========================================
# UTIL
# ==========================================
def pausa(seg=1):
    time.sleep(seg)

def mostrar_escritorio():
    pyautogui.hotkey("win", "d")
    pausa(1)

# ==========================================
# BUSCAR CRM
# ==========================================
def buscar_crm():
    for proc in psutil.process_iter(['pid', 'name']):
        try:
            if proc.info['name'] and proc.info['name'].lower() == "crm.exe":
                return proc
        except:
            continue
    return None

# ==========================================
# TRAER CRM AL FRENTE
# ==========================================
def traer_crm_al_frente():
    ventanas = pyautogui.getAllWindows()
    for ventana in ventanas:
        if ventana.title and "CRM" in ventana.title.upper():
            ventana.restore()
            ventana.activate()
            pausa(1)
            print("🟢 CRM traído al frente")
            return True
    print("⚠ No se pudo traer CRM al frente")
    return False

# ==========================================
# ABRIR CRM SI NO EXISTE
# ==========================================
def abrir_crm_si_no_existe():
    global crm_process
    existente = buscar_crm()
    if existente:
        crm_process = existente
        traer_crm_al_frente()
        print("🟢 CRM ya estaba abierto")
        return False
    print("🚀 Abriendo CRM...")
    mostrar_escritorio()
    os.startfile(CRM_PATH)
    for _ in range(30):
        pausa(1)
        crm_process = buscar_crm()
        if crm_process:
            traer_crm_al_frente()
            print(f"✅ CRM abierto PID {crm_process.pid}")
            return True
    print("❌ No se pudo abrir CRM")
    return False

# ==========================================
# ESPERAR IMAGEN CON F2
# ==========================================
def esperar_imagen_con_f2():
    print("🔍 Buscando imagen consultas...")
    reintentos = 0
    max_reintentos = 5
    while True:
        traer_crm_al_frente()
        try:
            location = pyautogui.locateOnScreen(IMAGEN_CONSULTAS, confidence=CONFIDENCE)
        except Exception:
            location = None
        if location:
            print("✅ Imagen detectada")
            return True
        pyautogui.press("f2")
        pausa(2)
        reintentos += 1
        if reintentos >= max_reintentos:
            print("⚠ No se detectó la imagen después de 5 intentos, volviendo a intentar...")
            reintentos = 0

# ==========================================
# ESPERAR IMAGEN DEL CRM ANTES DE LOGIN
# ==========================================
def esperar_imagen_crm():
    print("🔍 Esperando que aparezca la imagen del CRM...")
    while True:
        try:
            location = pyautogui.locateOnScreen(IMAGEN_CRM, confidence=CONFIDENCE)
        except Exception:
            location = None
        if location:
            print("✅ Imagen del CRM detectada, continuando flujo")
            return True
        pausa(2)

# ==========================================
# EXCEL: ENCONTRAR SIGUIENTE FILA PENDIENTE (solo TIPO = "UM")
# ==========================================
def obtener_siguiente_fila_pendiente():
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb.active

        col_estado = None
        col_otp    = None
        col_tipo   = None

        for idx, cell in enumerate(ws[1], 1):
            if cell.value == ESTADO_COLUMNA:
                col_estado = idx
            if cell.value == "OTP":
                col_otp = idx
            if cell.value == "TIPO":
                col_tipo = idx

        if not col_estado or not col_otp or not col_tipo:
            return None

        for row in range(2, ws.max_row + 1):
            estado   = ws.cell(row=row, column=col_estado).value
            otp_val  = ws.cell(row=row, column=col_otp).value
            tipo_val = ws.cell(row=row, column=col_tipo).value

            pendiente = (estado is None or str(estado).strip() == "")
            tiene_otp = (otp_val is not None and str(otp_val).strip() != "")
            es_um     = (tipo_val is not None and str(tipo_val).strip().upper() == "UM")

            if pendiente and tiene_otp and es_um:
                return row
        return None
    except:
        return None

# ==========================================
# EXCEL: OBTENER OTP DE UNA FILA
# ==========================================
def obtener_otp_de_fila(row):
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb.active

        col_otp = None
        for idx, cell in enumerate(ws[1], 1):
            if cell.value == "OTP":
                col_otp = idx
                break
        if not col_otp:
            return None

        val = ws.cell(row=row, column=col_otp).value
        if val is None:
            return None
        val = str(val).strip()
        if val == "":
            return None
        return val
    except:
        return None

# ==========================================
# EXCEL: MARCAR FILA COMO COMPLETADO/ERROR
# ==========================================
def marcar_fila_estado(row, valor):
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb.active

        col_idx = None
        for idx, cell in enumerate(ws[1], 1):
            if cell.value == ESTADO_COLUMNA:
                col_idx = idx
                break

        if not col_idx:
            print(f"⚠ No se encontró columna {ESTADO_COLUMNA} en Excel")
            return

        ws.cell(row=row, column=col_idx).value = valor
        wb.save(EXCEL_PATH)
        print(f"✅ Excel fila {row} marcada como {valor}")
    except Exception as e:
        print(f"⚠ Error marcando Excel: {e}")

# ==========================================
# COPIAR OTP (DE LA FILA ACTUAL) Y PEGAR
# ==========================================
def copiar_pegar_otp():
    global fila_actual_excel
    print("📋 Copiando OTP desde Excel y pegando...")
    try:
        if not fila_actual_excel:
            print("⚠ No hay fila_actual_excel")
            return False

        otp = obtener_otp_de_fila(fila_actual_excel)
        if not otp:
            print("⚠ OTP vacío o no encontrado en la fila actual")
            return False

        pyperclip.copy(otp)
        pausa(0.3)
        pyautogui.hotkey("ctrl", "v")
        pausa(0.3)
        pyautogui.press("enter")
        pausa(5)

        try:
            loc_ocupada = pyautogui.locateOnScreen(IMAGEN_OTP_OCUPADA, confidence=CONFIDENCE)
        except Exception:
            loc_ocupada = None

        if loc_ocupada:
            print("⚠ OTP ocupada detectada → presionando Enter")
            pyautogui.press("enter")
            pausa(1)
        else:
            print("✅ OTP libre, continuando flujo normal")

        print("✅ OTP pegado y Enter ejecutado")
        return True
    except Exception as e:
        print(f"⚠ Error en copiar_pegar_otp: {e}")
        return False

# ==========================================
# DETECTAR FLUJO Y MAXIMIZAR
# ==========================================
def detectar_flujo_y_maximizar():
    print("🔍 Buscando imagen flujo de trabajo...")
    reintentos = 0
    max_reintentos = 5
    while True:
        try:
            location = pyautogui.locateOnScreen(IMAGEN_FLUJO_TRABAJO, confidence=0.87)
        except Exception:
            location = None
        if location:
            print("✅ Flujo detectado → Maximizando ventana")
            pyautogui.hotkey("win", "up")
            pausa(2)
            return True
        pausa(2)
        reintentos += 1
        if reintentos >= max_reintentos:
            print("⚠ No se detectó el flujo después de 5 intentos, volviendo a intentar...")
            reintentos = 0

# ==========================================
# EXCEL: OBTENER DOCUMENTACION DE LA FILA ACTUAL
# ==========================================
def obtener_documentacion_de_excel():
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb.active

        col_doc = None
        for idx, cell in enumerate(ws[1], 1):
            if cell.value == "DOCUMENTACION":
                col_doc = idx
                break

        if not col_doc:
            print("⚠ No se encontró columna DOCUMENTACION en Excel")
            return None

        val = ws.cell(row=fila_actual_excel, column=col_doc).value
        if val is None:
            return None
        return str(val).strip().upper()
    except Exception as e:
        print(f"⚠ Error leyendo DOCUMENTACION: {e}")
        return None

# ==========================================
# EXCEL: OBTENER DOCUMENTACION_UM (texto a pegar en anotaciones)
# ==========================================
def obtener_documentacion_um_de_excel():
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb.active

        col_doc = None
        for idx, cell in enumerate(ws[1], 1):
            if cell.value == "DOCUMENTACION_UM":
                col_doc = idx
                break

        if not col_doc:
            return None

        val = ws.cell(row=fila_actual_excel, column=col_doc).value
        return str(val).strip() if val else None
    except Exception as e:
        print(f"⚠ Error leyendo DOCUMENTACION_UM: {e}")
        return None

# ==========================================
# GUARDAR EN CRM
# ==========================================
def guardar_crm():
    print("💾 Buscando botón guardar...")
    try:
        location_guardar = pyautogui.locateOnScreen(IMAGEN_GUARDAR, confidence=CONFIDENCE)
        if location_guardar:
            pyautogui.click(location_guardar)
            pausa(2)
            print("✅ Guardado correctamente")
            pyautogui.press("enter")
            pausa(1)
            pyautogui.press("enter")
            pausa(1)
            return True
        else:
            print("⚠ No se encontró guardar.png")
            return False
    except Exception as e:
        print(f"⚠ Error en guardar_crm: {e}")
        return False

# ==========================================
# TABS + BAJAR + PONER 0 (para editar tarea UM)
# ==========================================
def tabs_y_bajar_y_poner_cero_um():
    pyautogui.press("enter")
    pausa(1)
    pyautogui.press("tab")
    pausa(0.5)
    pyautogui.press("tab")
    pausa(0.2)
    for _ in range(4):
        pyautogui.press("tab")
        pausa(0.15)
    pyautogui.press("down")
    pausa(0.2)
    pyautogui.write("0", interval=0.05)
    pausa(0.2)
    pyautogui.press("tab")
    pausa(0.2)

# ==========================================
# CODIGOS DE RESOLUCION (UM)
# ==========================================
def ejecutar_codigos_de_resolucion_um():
    try:
        loc_codigos = pyautogui.locateOnScreen(IMAGEN_CODIGOS_RESOLUCION, confidence=CONFIDENCE)
    except Exception:
        loc_codigos = None

    if not loc_codigos:
        print("⚠ No se encontró codigos_de_resolucion.png")
        return False

    pyautogui.click(loc_codigos)
    pausa(0.6)

    try:
        loc = pyautogui.locateOnScreen(IMAGEN_COD_RESOLUCION_1, confidence=CONFIDENCE)
    except Exception:
        loc = None

    if not loc:
        print("⚠ No se encontró codigo_de_resolucion_1.png")
        return False

    x, y = pyautogui.center(loc)
    pyautogui.click(x + 20, y)
    pausa(0.4)
    pyautogui.press("tab")
    pausa(0.35)
    return True

# ==========================================
# ABRIR ITEM OTH + PREPARAR (doble click, editar, tabs, codigos, asignacion)
# ==========================================
def abrir_item_y_preparar_um(location_item):
    if not location_item:
        return False

    pyautogui.click(location_item)
    pyautogui.click(location_item)
    pausa(3)

    pyautogui.hotkey("win", "up")
    pausa(1)

    try:
        loc_editar = pyautogui.locateOnScreen(IMAGEN_EDITAR_TAREA, confidence=CONFIDENCE)
    except Exception:
        loc_editar = None

    if not loc_editar:
        print("⚠ No se encontró editar_tarea.png")
        return False

    pyautogui.click(loc_editar)
    pausa(1)

    try:
        loc_asig = pyautogui.locateOnScreen(IMAGEN_ASIGNACION_TAREA, confidence=CONFIDENCE)
    except Exception:
        loc_asig = None
    if loc_asig:
        pyautogui.click(loc_asig)
        pyautogui.press("enter")
        pausa(5)
    else:
        print("⚠ No se encontró asignacion_tarea.png")

    return True

# ==========================================
# PEGAR TEXTO EN ANOTACIONES (CON LIMITE 900 + REINGRESO)
# ==========================================
def pegar_texto_anotaciones_um(texto, location_item):
    if texto is None or str(texto).strip() == "":
        return True

    x_item, y_item = pyautogui.center(location_item)
    print(f"📍 Coordenadas item original: x={x_item}, y={y_item}")

    partes = dividir_texto_por_limite(texto)

    for i, parte in enumerate(partes):
        try:
            location_anotaciones = pyautogui.locateOnScreen(IMAGEN_ANOTACIONES, confidence=CONFIDENCE)
        except Exception:
            location_anotaciones = None

        if not location_anotaciones:
            print("⚠ No se encontró anotaciones.png")
            return False

        x, y = pyautogui.center(location_anotaciones)
        pyautogui.moveTo(x + 20, y + 10)
        pyautogui.click()
        pausa(1)

        parte = str(parte).replace('\r\n', '\n').replace('\r', '\n')
        lineas = parte.split('\n')

        for j, linea in enumerate(lineas):
            if linea:
                win32clipboard.OpenClipboard()
                win32clipboard.EmptyClipboard()
                win32clipboard.SetClipboardText(linea, win32clipboard.CF_UNICODETEXT)
                win32clipboard.CloseClipboard()
                pyautogui.hotkey("ctrl", "v")
                pausa(0.1)
            if j < len(lineas) - 1:
                pyautogui.press("enter")
                pausa(0.1)

        pausa(0.3)
        pyautogui.press("enter")

        if i < len(partes) - 1:
            print("💾 Límite 900 alcanzado, guardando y reingresando...")

            if not guardar_crm():
                return False

            try:
                loc_cierre = pyautogui.locateOnScreen(IMAGEN_CIERRE, confidence=CONFIDENCE)
            except Exception:
                loc_cierre = None
            if not loc_cierre:
                print("⚠ No se encontró cierre.png durante reintento por límite")
                return False
            pyautogui.click(loc_cierre)
            pausa(4)

            print(f"🔁 Reingresando con coordenadas originales: x={x_item}, y={y_item}")
            pyautogui.click(x_item, y_item)
            pyautogui.click(x_item, y_item)
            pausa(5)

            pyautogui.hotkey("win", "up")
            pausa(1)

            try:
                loc_editar = pyautogui.locateOnScreen(IMAGEN_EDITAR_TAREA, confidence=CONFIDENCE)
                if loc_editar:
                    pyautogui.click(loc_editar)
                    pausa(0.7)
                    ejecutar_codigos_de_resolucion_um()
            except Exception:
                print("⚠ No se encontró editar_tarea.png en reingreso")

            pausa(2)

    print("✅ Documentación UM pegada en anotaciones")
    return True

# ==========================================
# PROCESAR TAREAS UM
# ==========================================
def procesar_tareas_um():
    tipo_doc = obtener_documentacion_de_excel()
    print(f"📋 DOCUMENTACION leída del Excel: {tipo_doc}")

    MAPA_IMAGENES = {
        "OTH VISITA TECNICA":         IMAGEN_VISITA_TECNICA,
        "OTH INSTALACION DE SERVICIO": IMAGEN_OTH_INSTALACION,
        "OTH ACOMETIDA":              IMAGEN_OTH_ACOMETIDA,
        "OTH DISEÑO UM":              IMAGEN_OTH_DISEÑO_UM,
    }

    # ---- CASO OTP ----
    if tipo_doc == "OTP":
        print("🔀 Tipo OTP → flujo especial sin TAREAS")

        pyautogui.press("enter")
        pausa(2)
        pyautogui.press("enter")

        try:
            loc_editar = pyautogui.locateOnScreen(IMAGEN_EDITAR_INCIDENTE, confidence=CONFIDENCE)
        except Exception:
            loc_editar = None

        if not loc_editar:
            print("⚠ No se encontró editar_incidente.png")
            return False

        pyautogui.click(loc_editar)
        pausa(1)
        pyautogui.press("enter")

        texto_um = obtener_documentacion_um_de_excel()
        if not pegar_texto_anotaciones_um(texto_um, loc_editar):
            return False

        try:
            loc_guardar_otp = pyautogui.locateOnScreen(IMAGEN_GUARDAR_OTP, confidence=CONFIDENCE)
        except Exception:
            loc_guardar_otp = None

        if not loc_guardar_otp:
            print("⚠ No se encontró guardar_otp.png")
            return False

        pyautogui.click(loc_guardar_otp)
        pausa(2)

        try:
            loc_cierre = pyautogui.locateOnScreen(IMAGEN_CIERRE, confidence=CONFIDENCE)
        except Exception:
            loc_cierre = None

        if loc_cierre:
            pyautogui.click(loc_cierre)
            pausa(2)
        else:
            print("⚠ No se encontró cierre.png")

        return True

    # ---- CASOS OTH ----
    imagen_oth = MAPA_IMAGENES.get(tipo_doc)
    if not imagen_oth:
        print(f"⚠ Tipo desconocido: {tipo_doc}")
        return False

    try:
        location_tareas = pyautogui.locateOnScreen(IMAGEN_TAREAS, confidence=CONFIDENCE)
    except Exception:
        location_tareas = None

    if not location_tareas:
        print("⚠ No se detectó tareas.png")
        return False

    pyautogui.click(location_tareas)
    pausa(1)
    pyautogui.press("enter")
    pausa(1)
    pyautogui.move(0, 30)
    pausa(1)

    pyautogui.scroll(-20)
    pausa(2)

    loc_oth = None
    for _ in range(3):
        try:
            loc_oth = pyautogui.locateOnScreen(imagen_oth, confidence=CONFIDENCE)
        except Exception:
            loc_oth = None

        if loc_oth:
            print(f"✅ Encontrado {tipo_doc}")
            break

        print(f"🔍 No se encontró {tipo_doc}, haciendo scroll y reintentando...")
        pyautogui.scroll(-10)
        pausa(3)

    if not loc_oth:
        print(f"⚠ No se encontró imagen OTH tras scrolls: {imagen_oth}")
        return False

    pyautogui.scroll(-3)
    pausa(3)

    try:
        loc_oth = pyautogui.locateOnScreen(imagen_oth, confidence=CONFIDENCE)
    except Exception:
        loc_oth = None

    if not loc_oth:
        print(f"⚠ No se encontró imagen OTH tras segundo scroll: {imagen_oth}")
        return False

    if not abrir_item_y_preparar_um(loc_oth):
        return False

    texto_um = obtener_documentacion_um_de_excel()
    if not pegar_texto_anotaciones_um(texto_um, loc_oth):
        return False

    if not guardar_crm():
        return False

    return True

# ==========================================
# FLUJO PRINCIPAL UM (hasta ok3)
# ==========================================
def ejecutar_flujo(necesita_login):
    global flujo_en_ejecucion, last_activity_time

    with lock:
        if flujo_en_ejecucion:
            return
        flujo_en_ejecucion = True

    traer_crm_al_frente()
    print("⚡ Ejecutando flujo UM...")
    pausa(2)

    if necesita_login:
        user = os.environ.get("CRM_USER", "46381573").strip()
        pwd  = os.environ.get("CRM_PASS",  "*Dairy136*").strip()

        if not user or not pwd:
            print("❌ Faltan credenciales. Configura CRM_USER y CRM_PASS en variables de entorno.")
            with lock:
                flujo_en_ejecucion = False
            return

        esperar_imagen_crm()
        pyautogui.press("tab"); pausa(1)
        pyautogui.press("tab"); pausa(1)
        pyautogui.write(user, interval=0.05); pausa(1)
        pyautogui.press("tab"); pausa(1)
        pyautogui.write(pwd,  interval=0.05); pausa(1)
        pyautogui.press("tab"); pausa(1)
        pyautogui.press("enter"); pausa(3)
        pyautogui.press("enter"); pausa(2)
        for _ in range(4):
            pyautogui.press("tab"); pausa(1)
        pyautogui.press("enter"); pausa(15)
        pyautogui.press("f2"); pausa(2)

    ok1 = esperar_imagen_con_f2()
    ok2 = copiar_pegar_otp()
    ok3 = detectar_flujo_y_maximizar()
    ok4 = procesar_tareas_um()

    if ok1 and ok2 and ok3 and ok4:
        marcar_fila_estado(fila_actual_excel, "COMPLETADO")
    else:
        marcar_fila_estado(fila_actual_excel, "ERROR")

    last_activity_time = time.time()
    with lock:
        flujo_en_ejecucion = False

    print("🏁 Flujo UM terminado")

# ==========================================
# INICIO (COLA POR EXCEL - solo filas UM)
# ==========================================
if __name__ == "__main__":
    if not adquirir_lock():
        print("🟡 Ya hay un bot UM corriendo. Este se cierra para respetar la cola por Excel.")
    else:
        try:
            necesita_login_global = abrir_crm_si_no_existe()
            inicio_espera_sin_pendientes = None

            while True:
                time.sleep(1)

                if flujo_en_ejecucion:
                    inicio_espera_sin_pendientes = None
                    continue

                siguiente = obtener_siguiente_fila_pendiente()

                if siguiente:
                    inicio_espera_sin_pendientes = None
                    fila_actual_excel = siguiente
                    adquirir_cola_lock()
                    try:
                        ejecutar_flujo(necesita_login_global)
                    finally:
                        liberar_cola_lock()
                    necesita_login_global = False
                    continue

                if inicio_espera_sin_pendientes is None:
                    inicio_espera_sin_pendientes = time.time()
                    print("⏳ Sin pendientes UM. Iniciando espera de 8 minutos...")

                if time.time() - inicio_espera_sin_pendientes >= INACTIVITY_LIMIT:
                    print("🔴 8 minutos sin actividad. Cerrando CRM...")
                    try:
                        if crm_process:
                            crm_process.terminate()
                        else:
                            proc = buscar_crm()
                            if proc:
                                proc.terminate()
                        print("✅ CRM cerrado por inactividad")
                    except Exception as e:
                        print(f"⚠ Error cerrando CRM: {e}")
                    break

        finally:
            liberar_lock()

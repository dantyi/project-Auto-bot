from flask import Flask, render_template, request, jsonify
from werkzeug.utils import secure_filename
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
import os
import subprocess

app = Flask(__name__)
EXCEL_FILE = "datos.xlsx"
OLE_FOLDER = r"C:\Users\dell\Desktop\kickoff\project-Auto-bot\OLE"

# ==========================
# CREAR EXCEL SI NO EXISTE
# ==========================
def crear_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Registros"

        # ✅ ESTRUCTURA ORIGINAL (NO SE TOCA) + ADICIONALES AL FINAL
        encabezados = [
            "OTP",
            "DOCUMENTAR_CHECK_FACTIBILIDAD",
            "CORREO_REPORTE_INICIO",
            "MARCACION_OTH",
            "COD_RESOLUCION_1",
            "GERENCIA",
            "FECHA_COMPROMISO",
            "FECHA_PROGRAMACION",
            "TIPO_SERVICIO",
            "COMPLETADO",
            "DOCUMENTACION_ITEM_FACTURACION",
            "CERRADO_OTP",
            "COD_RESOLUCION_1_OTP",
            "MODIFICAR_OTP",
            "TIPO",
            "DOCUMENTACION_UM",
            "CERRAR_OTH",
            "DOCUMENTACION",
            "DOC_CONFIG",
            "OT_TIPO",
            "PATH_PRUEBAS_PREVIAS",
            "PATH_MINUTOGRAMA",
            "PATH_VALIDACION_WAN_LAN",
            "PATH_SCRIPT",
            "PATH_SATURACION"
        ]

        ws.append(encabezados)
        wb.save(EXCEL_FILE)

def agregar_columnas_faltantes():
    """Agrega columnas nuevas al Excel si ya existe pero le faltan."""
    nuevas = ["TIPO", "DOCUMENTACION_UM", "CERRAR_OTH", "DOCUMENTACION",
              "DOC_CONFIG", "OT_TIPO",
              "PATH_PRUEBAS_PREVIAS", "PATH_MINUTOGRAMA",
              "PATH_VALIDACION_WAN_LAN", "PATH_SCRIPT", "PATH_SATURACION"]
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    existentes = [cell.value for cell in ws[1]]
    agregado = False
    for col in nuevas:
        if col not in existentes:
            ws.cell(row=1, column=len(existentes) + 1, value=col)
            existentes.append(col)
            agregado = True
    if agregado:
        wb.save(EXCEL_FILE)

crear_excel()
agregar_columnas_faltantes()

# ==========================
# RUTAS PRINCIPALES
# ==========================
@app.route("/")
def index():
    return render_template("index.html")

@app.route("/um")
def um():
    return render_template("UM.html")

@app.route("/configuracion")
def configuracion():
    return render_template("configuracion.html")

# ==========================
# GUARDAR REGISTRO + ACTIVAR BOT
# ==========================
@app.route("/guardar", methods=["POST"])
def guardar():
    try:
        data = request.json
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active

        # ✅ FILA ORIGINAL (NO SE TOCA) + ADICIONALES AL FINAL
        fila = [
            data.get("otp", ""),
            data.get("factibilidad", ""),
            data.get("correo_inicio", ""),
            data.get("marcacion_oth", ""),
            data.get("cod_resolucion", ""),
            data.get("gerencia", ""),
            data.get("fecha_compromiso", ""),
            data.get("fecha_programacion", ""),
            data.get("tipo_servicio", ""),
            "",

            data.get("documentacion_item_facturacion", ""),  # "SI" / "NO"
            data.get("cerrado_otp", ""),                     # "SI" / "NO"
            data.get("cod_resolucion_otp", ""),              # TEXTO
            data.get("modificar_otp", ""),                   # "SI" / "NO"
            "KICKOFF",                                       # TIPO
            "",                                              # DOCUMENTACION_UM
            "",                                              # CERRAR_OTH
            ""                                               # DOCUMENTACION
        ]

        ws.append(fila)

        nueva_fila_num = ws.max_row
        for col in range(1, 19):  # 1..18
            ws.cell(row=nueva_fila_num, column=col).alignment = Alignment(wrap_text=True)

        wb.save(EXCEL_FILE)

        subprocess.Popen(["python", "bot.py"])

        return jsonify({"mensaje": "Guardado y BOT ejecutado correctamente"})

    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ==========================
# GUARDAR UM + ACTIVAR BOT UM
# ==========================
@app.route("/guardar_um", methods=["POST"])
def guardar_um():
    try:
        data = request.json
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active

        fila = [
            data.get("otp", ""),   # OTP
            "", "", "", "", "",    # cols 2-6 vacías
            "", "", "",            # cols 7-9 vacías
            "",                    # COMPLETADO
            "", "", "", "",        # cols 11-14 vacías
            "UM",                  # TIPO
            data.get("documentacion_um", ""),  # DOCUMENTACION_UM
            data.get("cerrar_oth", ""),         # CERRAR_OTH
            data.get("documentacion", "")       # DOCUMENTACION (OTH tipo)
        ]

        ws.append(fila)

        nueva_fila_num = ws.max_row
        for col in range(1, 19):  # 1..18
            ws.cell(row=nueva_fila_num, column=col).alignment = Alignment(wrap_text=True)

        wb.save(EXCEL_FILE)

        subprocess.Popen(["python", "bot_um.py"])

        return jsonify({"mensaje": "Guardado y BOT UM ejecutado correctamente"})

    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ==========================
# GUARDAR CONFIG + ACTIVAR BOT CONFIG
# ==========================
@app.route("/guardar_config", methods=["POST"])
def guardar_config():
    try:
        otp        = request.form.get("otp", "").strip()
        ot_tipo    = request.form.get("ot_tipo", "").strip()
        doc_config = request.form.get("doc_config", "").strip()

        os.makedirs(OLE_FOLDER, exist_ok=True)

        def guardar_seccion(campo):
            archivos = request.files.getlist(campo)
            rutas = []
            for archivo in archivos:
                if archivo and archivo.filename:
                    nombre = secure_filename(archivo.filename).lower()
                    nombre_nuevo = f"{otp}_{nombre}"
                    ruta = os.path.join(OLE_FOLDER, nombre_nuevo)
                    archivo.save(ruta)
                    rutas.append(ruta)
            return "; ".join(rutas) if rutas else "NO"

        path_pruebas   = guardar_seccion("archivos_pruebas_previas")
        path_minuto    = guardar_seccion("archivos_minutograma")
        path_wanlan    = guardar_seccion("archivos_wanlan")
        path_script    = guardar_seccion("archivos_script")
        path_saturacion = guardar_seccion("archivos_saturacion")

        wb = load_workbook(EXCEL_FILE)
        ws = wb.active

        fila = [
            otp,             # OTP
            "", "", "", "", "",    # cols 2-6
            "", "", "",           # cols 7-9
            "",                   # COMPLETADO
            "", "", "", "",       # cols 11-14
            "CONFIGURACION",      # TIPO
            "",                   # DOCUMENTACION_UM
            "",                   # CERRAR_OTH
            doc_config,           # DOCUMENTACION
            "",                   # DOC_CONFIG
            ot_tipo,              # OT_TIPO
            path_pruebas,         # PATH_PRUEBAS_PREVIAS
            path_minuto,          # PATH_MINUTOGRAMA
            path_wanlan,          # PATH_VALIDACION_WAN_LAN
            path_script,          # PATH_SCRIPT
            path_saturacion       # PATH_SATURACION
        ]

        ws.append(fila)

        nueva_fila_num = ws.max_row
        for col in range(1, 26):
            ws.cell(row=nueva_fila_num, column=col).alignment = Alignment(wrap_text=True)

        wb.save(EXCEL_FILE)

        subprocess.Popen(["python", "bot_configuracion.py"])

        return jsonify({"mensaje": "Guardado y BOT Configuración ejecutado correctamente"})

    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ==========================
# EJECUTAR EN RED 0.0.0.0
# ==========================
if __name__ == "__main__":
    app.run(
        host="0.0.0.0",
        port=5000,
        debug=False
    )
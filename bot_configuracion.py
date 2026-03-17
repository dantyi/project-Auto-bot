import subprocess
import psutil
import threading
import time
import pyautogui
import os
import openpyxl
import pyperclip
import win32clipboard

# ==========================================
# CONFIG
# ==========================================
CRM_PATH             = r"C:\Users\dell\AppData\Roaming\Microsoft\Windows\Start Menu\Programs\Global Hitss\Global Hitss\Nuevo CRM.appref-ms"
IMAGEN_CONSULTAS     = r"C:\Users\dell\Desktop\kickoff\project-Auto-bot\stratic\consultas.png"
IMAGEN_FLUJO_TRABAJO = r"C:\Users\dell\Desktop\kickoff\project-Auto-bot\stratic\flujo de trabajo.png"
IMAGEN_CRM           = r"C:\Users\dell\Desktop\kickoff\project-Auto-bot\stratic\CRM.png"
IMAGEN_OTP_OCUPADA   = r"C:\Users\dell\Desktop\kickoff\project-Auto-bot\stratic\OTPOCUPADA.png"
IMAGEN_ANOTACIONES     = r"C:\Users\dell\Desktop\kickoff\project-Auto-bot\stratic\anotaciones.png"
IMAGEN_GUARDAR         = r"C:\Users\dell\Desktop\kickoff\project-Auto-bot\stratic\guardar.png"
IMAGEN_CIERRE          = r"C:\Users\dell\Desktop\kickoff\project-Auto-bot\stratic\cierre.png"
IMAGEN_TAREAS              = r"C:\Users\dell\Desktop\kickoff\project-Auto-bot\stratic\tareas.png"
IMAGEN_TAREAS_1            = r"C:\Users\dell\Desktop\kickoff\project-Auto-bot\stratic\tareas_1.png"
IMAGEN_CONFIGURACION       = r"C:\Users\dell\Desktop\kickoff\project-Auto-bot\stratic\configuracion.png"
IMAGEN_EDITAR_TAREA        = r"C:\Users\dell\Desktop\kickoff\project-Auto-bot\stratic\editar_tarea.png"
IMAGEN_FECHA_COMPROMISO    = r"C:\Users\dell\Desktop\kickoff\project-Auto-bot\stratic\FECHA_compromiso.png"
IMAGEN_FECHA_PROGRAMACION  = r"C:\Users\dell\Desktop\kickoff\project-Auto-bot\stratic\fecha_programacion1.png"
IMAGEN_TIPO_SERVICIO       = r"C:\Users\dell\Desktop\kickoff\project-Auto-bot\stratic\tipo_de_servicio.png"
IMAGEN_OLE               = r"C:\Users\dell\Desktop\kickoff\project-Auto-bot\stratic\OLE.png"
IMAGEN_EDITAR_INCIDENTE  = r"C:\Users\dell\Desktop\kickoff\project-Auto-bot\stratic\editar_incidente.png"
IMAGEN_ASIGNAR_INCIDENTE = r"C:\Users\dell\Desktop\kickoff\project-Auto-bot\stratic\asignar_incidente.png"
IMAGEN_AGREGAR           = r"C:\Users\dell\Desktop\kickoff\project-Auto-bot\stratic\agregar.png"
IMAGEN_ACEPTAR           = r"C:\Users\dell\Desktop\kickoff\project-Auto-bot\stratic\aceptar.png"

EXCEL_PATH     = r"C:\Users\dell\Desktop\kickoff\project-Auto-bot\datos.xlsx"
ESTADO_COLUMNA = "COMPLETADO"
INACTIVITY_LIMIT = 480
CONFIDENCE = 0.87

crm_process = None
last_activity_time = time.time()
flujo_en_ejecucion = False
lock = threading.Lock()
fila_actual_excel = None

pyautogui.FAILSAFE = True

MAX_CARACTERES_CRM = 1000
ENTER_COST = 4

# ==========================================
# DIVIDIR TEXTO POR LIMITE
# ==========================================
def dividir_texto_por_limite(texto, limite=MAX_CARACTERES_CRM):
    if texto is None:
        return []
    texto = str(texto).replace("\r\n", "\n").replace("\r", "\n")
    partes = []
    actual = []
    costo_actual = 0

    for linea in texto.split("\n"):
        c_linea = len(linea)
        c_enter = ENTER_COST if len(actual) > 0 else 0

        if (costo_actual + c_enter + c_linea) > limite and len(actual) > 0:
            partes.append("\n".join(actual))
            actual = []
            costo_actual = 0
            c_enter = 0

        if c_linea > limite:
            if len(actual) > 0:
                partes.append("\n".join(actual))
                actual = []
                costo_actual = 0
            start = 0
            while start < len(linea):
                end = min(start + limite, len(linea))
                partes.append(linea[start:end])
                start = end
            continue

        if c_enter:
            costo_actual += c_enter
        actual.append(linea)
        costo_actual += c_linea

    if len(actual) > 0:
        partes.append("\n".join(actual))

    return partes

# ====== LOCK PROPIO ======
LOCK_FILE = os.path.join(os.path.dirname(__file__), "bot_configuracion.lock")
lock_fd = None

def adquirir_lock():
    global lock_fd
    try:
        lock_fd = os.open(LOCK_FILE, os.O_CREAT | os.O_EXCL | os.O_RDWR)
        os.write(lock_fd, str(os.getpid()).encode())
        return True
    except FileExistsError:
        return False

def liberar_lock():
    global lock_fd
    try:
        if lock_fd is not None:
            os.close(lock_fd)
            lock_fd = None
        if os.path.exists(LOCK_FILE):
            os.remove(LOCK_FILE)
    except:
        pass

# ====== LOCK DE COLA COMPARTIDA ======
COLA_LOCK_FILE = os.path.join(os.path.dirname(__file__), "cola.lock")
cola_lock_fd = None

def adquirir_cola_lock():
    global cola_lock_fd
    intentos = 0
    while True:
        try:
            cola_lock_fd = os.open(COLA_LOCK_FILE, os.O_CREAT | os.O_EXCL | os.O_RDWR)
            os.write(cola_lock_fd, str(os.getpid()).encode())
            return True
        except FileExistsError:
            if intentos % 10 == 0:
                print("⏳ Cola ocupada por otro bot, esperando...")
            time.sleep(1)
            intentos += 1

def liberar_cola_lock():
    global cola_lock_fd
    try:
        if cola_lock_fd is not None:
            os.close(cola_lock_fd)
            cola_lock_fd = None
        if os.path.exists(COLA_LOCK_FILE):
            os.remove(COLA_LOCK_FILE)
    except:
        pass

# ==========================================
# UTIL
# ==========================================
def pausa(seg=1):
    time.sleep(seg)

def mostrar_escritorio():
    pyautogui.hotkey("win", "d")
    pausa(1)

# ==========================================
# BUSCAR / ABRIR / TRAER CRM
# ==========================================
def buscar_crm():
    for proc in psutil.process_iter(['pid', 'name']):
        try:
            if proc.info['name'] and proc.info['name'].lower() == "crm.exe":
                return proc
        except:
            continue
    return None

def traer_crm_al_frente():
    ventanas = pyautogui.getAllWindows()
    for ventana in ventanas:
        if ventana.title and "CRM" in ventana.title.upper():
            ventana.restore()
            ventana.activate()
            pausa(1)
            print("🟢 CRM traído al frente")
            return True
    print("⚠ No se pudo traer CRM al frente")
    return False

def abrir_crm_si_no_existe():
    global crm_process
    existente = buscar_crm()
    if existente:
        crm_process = existente
        traer_crm_al_frente()
        print("🟢 CRM ya estaba abierto")
        return False
    print("🚀 Abriendo CRM...")
    mostrar_escritorio()
    os.startfile(CRM_PATH)
    for _ in range(30):
        pausa(1)
        crm_process = buscar_crm()
        if crm_process:
            traer_crm_al_frente()
            print(f"✅ CRM abierto PID {crm_process.pid}")
            return True
    print("❌ No se pudo abrir CRM")
    return False

def esperar_imagen_con_f2():
    print("🔍 Buscando imagen consultas...")
    reintentos = 0
    while True:
        traer_crm_al_frente()
        try:
            location = pyautogui.locateOnScreen(IMAGEN_CONSULTAS, confidence=CONFIDENCE)
        except Exception:
            location = None
        if location:
            print("✅ Imagen detectada")
            return True
        pyautogui.press("f2")
        pausa(2)
        reintentos += 1
        if reintentos >= 5:
            print("⚠ No se detectó la imagen después de 5 intentos, reintentando...")
            reintentos = 0

def esperar_imagen_crm():
    print("🔍 Esperando imagen del CRM...")
    while True:
        try:
            location = pyautogui.locateOnScreen(IMAGEN_CRM, confidence=CONFIDENCE)
        except Exception:
            location = None
        if location:
            print("✅ Imagen del CRM detectada")
            return True
        pausa(2)

# ==========================================
# EXCEL
# ==========================================
def obtener_siguiente_fila_pendiente():
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb.active

        col_estado = col_otp = col_tipo = None
        for idx, cell in enumerate(ws[1], 1):
            if cell.value == ESTADO_COLUMNA: col_estado = idx
            if cell.value == "OTP":          col_otp    = idx
            if cell.value == "TIPO":         col_tipo   = idx

        if not col_estado or not col_otp or not col_tipo:
            return None

        for row in range(2, ws.max_row + 1):
            estado   = ws.cell(row=row, column=col_estado).value
            otp_val  = ws.cell(row=row, column=col_otp).value
            tipo_val = ws.cell(row=row, column=col_tipo).value

            pendiente = (estado is None or str(estado).strip() == "")
            tiene_otp = (otp_val is not None and str(otp_val).strip() != "")
            es_config = (tipo_val is not None and str(tipo_val).strip().upper() == "CONFIGURACION")

            if pendiente and tiene_otp and es_config:
                return row
        return None
    except:
        return None

def obtener_otp_de_fila(row):
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb.active
        col_otp = None
        for idx, cell in enumerate(ws[1], 1):
            if cell.value == "OTP":
                col_otp = idx
                break
        if not col_otp:
            return None
        val = ws.cell(row=row, column=col_otp).value
        return str(val).strip() if val else None
    except:
        return None

def marcar_fila_estado(row, valor):
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb.active
        col_idx = None
        for idx, cell in enumerate(ws[1], 1):
            if cell.value == ESTADO_COLUMNA:
                col_idx = idx
                break
        if not col_idx:
            return
        ws.cell(row=row, column=col_idx).value = valor
        wb.save(EXCEL_PATH)
        print(f"✅ Excel fila {row} marcada como {valor}")
    except Exception as e:
        print(f"⚠ Error marcando Excel: {e}")

def obtener_dato_de_columna(nombre_col):
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb.active
        col_idx = None
        for idx, cell in enumerate(ws[1], 1):
            if cell.value == nombre_col:
                col_idx = idx
                break
        if not col_idx:
            return None
        val = ws.cell(row=fila_actual_excel, column=col_idx).value
        return str(val).strip() if val else None
    except Exception as e:
        print(f"⚠ Error leyendo {nombre_col}: {e}")
        return None

# ==========================================
# CRM: COPIAR / PEGAR OTP
# ==========================================
def copiar_pegar_otp():
    global fila_actual_excel
    print("📋 Copiando OTP desde Excel y pegando...")
    try:
        if not fila_actual_excel:
            return False
        otp = obtener_otp_de_fila(fila_actual_excel)
        if not otp:
            return False
        pyperclip.copy(otp)
        pausa(0.3)
        pyautogui.hotkey("ctrl", "v")
        pausa(0.3)
        pyautogui.press("enter")
        pausa(5)
        try:
            loc_ocupada = pyautogui.locateOnScreen(IMAGEN_OTP_OCUPADA, confidence=CONFIDENCE)
        except Exception:
            loc_ocupada = None
        if loc_ocupada:
            pyautogui.press("enter")
            pausa(1)
        print("✅ OTP pegado")
        return True
    except Exception as e:
        print(f"⚠ Error en copiar_pegar_otp: {e}")
        return False

def detectar_flujo_y_maximizar():
    print("🔍 Buscando flujo de trabajo...")
    reintentos = 0
    while True:
        try:
            location = pyautogui.locateOnScreen(IMAGEN_FLUJO_TRABAJO, confidence=0.87)
        except Exception:
            location = None
        if location:
            print("✅ Flujo detectado → Maximizando")
            pyautogui.hotkey("win", "up")
            pausa(2)
            return True
        pausa(2)
        reintentos += 1
        if reintentos >= 5:
            print("⚠ No se detectó el flujo, reintentando...")
            reintentos = 0

# ==========================================
# GUARDAR EN CRM
# ==========================================
def guardar_crm():
    print("💾 Guardando en CRM...")
    try:
        loc = pyautogui.locateOnScreen(IMAGEN_GUARDAR, confidence=CONFIDENCE)
        if loc:
            pyautogui.click(loc)
            pausa(2)
            pyautogui.press("enter")
            pausa(1)
            pyautogui.press("enter")
            pausa(1)
            print("✅ Guardado")
            return True
        print("⚠ No se encontró guardar.png")
        return False
    except Exception as e:
        print(f"⚠ Error guardando: {e}")
        return False

# ==========================================
# PEGAR TEXTO EN ANOTACIONES (CON LIMITE 900)
# ==========================================
def pegar_texto_en_anotaciones(texto, location_item):
    if texto is None or str(texto).strip() == "":
        return True

    x_item, y_item = pyautogui.center(location_item)
    print(f"📍 Coordenadas item original: x={x_item}, y={y_item}")

    partes = dividir_texto_por_limite(texto)

    for i, parte in enumerate(partes):
        try:
            loc_anot = pyautogui.locateOnScreen(IMAGEN_ANOTACIONES, confidence=CONFIDENCE)
        except Exception:
            loc_anot = None

        if not loc_anot:
            print("⚠ No se encontró anotaciones.png")
            return False

        x, y = pyautogui.center(loc_anot)
        pyautogui.click(x + 20, y + 10)
        pausa(1)

        parte = str(parte).replace('\r\n', '\n').replace('\r', '\n')
        lineas = parte.split('\n')

        for j, linea in enumerate(lineas):
            if linea:
                win32clipboard.OpenClipboard()
                win32clipboard.EmptyClipboard()
                win32clipboard.SetClipboardText(linea, win32clipboard.CF_UNICODETEXT)
                win32clipboard.CloseClipboard()
                pyautogui.hotkey("ctrl", "v")
                pausa(0.1)

            if j < len(lineas) - 1:
                pyautogui.press("enter")
                pausa(0.1)

        pausa(0.3)
        pyautogui.press("enter")

        if i < len(partes) - 1:
            print("💾 Límite 900 alcanzado, guardando y reingresando...")

            if not guardar_crm():
                return False

            try:
                loc_cierre = pyautogui.locateOnScreen(IMAGEN_CIERRE, confidence=CONFIDENCE)
            except Exception:
                loc_cierre = None
            if not loc_cierre:
                print("⚠ No se encontró cierre.png durante reintento por límite")
                return False
            pyautogui.click(loc_cierre)
            pausa(4)

            print(f"🔁 Reingresando con coordenadas originales: x={x_item}, y={y_item}")
            pyautogui.scroll(-2)
            pausa(2)
            pyautogui.click(x_item, y_item)
            pyautogui.click(x_item, y_item)
            pausa(5)

            pyautogui.hotkey("win", "up")
            pausa(1)

            try:
                loc_editar = pyautogui.locateOnScreen(IMAGEN_EDITAR_TAREA, confidence=CONFIDENCE)
                if loc_editar:
                    pyautogui.click(loc_editar)
                    pausa(1)
                    pyautogui.press("enter")
                    pausa(1)
            except:
                print("⚠ No se encontró editar_tarea.png en reingreso")

            pausa(2)

    print("✅ Texto pegado en anotaciones")
    return True

# ==========================================
# ITEMS DE FACTURACION (sin escribir fechas, solo ESC y navegar)
# ==========================================
def llenar_items_facturacion():

    # --- FECHA COMPROMISO ---
    try:
        loc = pyautogui.locateOnScreen(IMAGEN_FECHA_COMPROMISO, confidence=CONFIDENCE)
    except Exception:
        loc = None

    if not loc:
        print("⚠ No se encontró FECHA_compromiso.png")
        return False

    x, y = pyautogui.center(loc)
    pyautogui.click(x + 450, y)
    pausa(0.2)
    pyautogui.press("esc")
    pausa(0.2)

    pyautogui.press("tab");  pausa(0.2)
    pyautogui.press("down"); pausa(0.2)
    pyautogui.press("down"); pausa(0.2)
    pyautogui.write("inteegra"); pausa(0.2)

    pyautogui.press("tab"); pausa(0.2)
    pyautogui.write("NO")

    pyautogui.press("tab"); pausa(0.2)
    pyautogui.press("down"); pausa(0.2)
    pyautogui.press("down"); pausa(0.2)
    pyautogui.write("config"); pausa(0.2)

    pyautogui.press("tab"); pausa(0.2)


    pyautogui.scroll(-30); pausa(0.2)

    print("✅ FECHA_COMPROMISO navegada (ESC)")

    # --- OT_TIPO ---
    ot_tipo = obtener_dato_de_columna("OT_TIPO")
    print(f"📋 OT_TIPO: {ot_tipo}")

    if ot_tipo and "CORTA" in ot_tipo.upper():
        pausa(1)
        for _ in range(12):
            pyautogui.press("down"); pausa(0.1)

        pyautogui.press("tab"); pausa(0.2)
        pyautogui.write("30"); pausa(0.2)
        pyautogui.press("tab"); pausa(0.2)
        pyautogui.write("30"); pausa(0.2)

    else:
        pausa(1)
        for _ in range(13):
            pyautogui.press("down"); pausa(0.1)
            
        pyautogui.press("tab"); pausa(0.2)
        pyautogui.write("60"); pausa(0.2)
        pyautogui.press("tab"); pausa(0.2)
        pyautogui.write("60"); pausa(0.2)

    pyautogui.press("tab"); pausa(0.2)
    pyautogui.write("NO"); pausa(0.2)

    # 21 tabs → SI
    for _ in range(21):
        pyautogui.press("tab"); pausa(0.1)
    pyautogui.write("SI"); pausa(0.2)

    # 28 tabs → SI
    for _ in range(28):
        pyautogui.press("tab"); pausa(0.1)
    pyautogui.write("SI"); pausa(0.2)

    # 1 tab → 1
    pyautogui.press("tab"); pausa(0.2)
    pyautogui.write("1"); pausa(0.2)

    print("✅ Items de facturación completados")
    return True

# ==========================================
# OBTENER PATHS DEL EXCEL (solo los que no son "NO")
# ==========================================
def obtener_paths_de_excel():
    COLUMNAS = [
        ("PATH_PRUEBAS_PREVIAS",    "PRUEBAS PREVIAS"),
        ("PATH_MINUTOGRAMA",        "MINUTOGRAMA"),
        ("PATH_VALIDACION_WAN_LAN", "VALIDACION WAN LAN"),
        ("PATH_SCRIPT",             "SCRIPT"),
        ("PATH_SATURACION",         "SATURACION"),
    ]
    resultado = []
    for col, etiqueta in COLUMNAS:
        val = obtener_dato_de_columna(col)
        if val and val.upper() != "NO":
            for ruta in val.split(";"):
                ruta = ruta.strip()
                if ruta:
                    resultado.append((ruta, etiqueta))
    return resultado

# ==========================================
# PROCESAR ADJUNTOS OLE (antes de TAREAS)
# ==========================================
def procesar_adjuntos_ole(paths):
    # 1. Buscar OLE.png (90%) → Enter
    try:
        loc_ole = pyautogui.locateOnScreen(IMAGEN_OLE, confidence=0.90)
    except Exception:
        loc_ole = None
    if not loc_ole:
        print("⚠ No se encontró OLE.png → saltando adjuntos")
        return False
    pyautogui.click(loc_ole)
    pausa(1)
    pyautogui.press("enter")
    pausa(2)

    # 2. Buscar editar_incidente.png → click
    try:
        loc_editar = pyautogui.locateOnScreen(IMAGEN_EDITAR_INCIDENTE, confidence=CONFIDENCE)
    except Exception:
        loc_editar = None
    if not loc_editar:
        print("⚠ No se encontró editar_incidente.png → saltando adjuntos")
        return False
    pyautogui.click(loc_editar)
    pausa(1)

    # 4. Buscar asignar_incidente.png → click
    try:
        loc_asig = pyautogui.locateOnScreen(IMAGEN_ASIGNAR_INCIDENTE, confidence=CONFIDENCE)
    except Exception:
        loc_asig = None
    if not loc_asig:
        print("⚠ No se encontró asignar_incidente.png → saltando adjuntos")
        return False
    pyautogui.click(loc_asig)
    pausa(1)

    # 5. Buscar asignar_incidente.png de nuevo → si no aparece sigue flujo normal
    try:
        loc_asig2 = pyautogui.locateOnScreen(IMAGEN_ASIGNAR_INCIDENTE, confidence=CONFIDENCE)
    except Exception:
        loc_asig2 = None
    if not loc_asig2:
        print("⚠ asignar_incidente.png no reaparece → continuando con flujo TAREAS")
        return False
    pyautogui.press("enter")
    pausa(2)

    # 6. Buscar agregar.png → Enter
    try:
        loc_agregar = pyautogui.locateOnScreen(IMAGEN_AGREGAR, confidence=CONFIDENCE)
    except Exception:
        loc_agregar = None
    if not loc_agregar:
        print("⚠ No se encontró agregar.png")
        return False
    pyautogui.click(loc_agregar)
    pausa(3)
    pyautogui.press("enter")
    pausa(5)

    # 7. Por cada archivo: (2do+ → Enter abre diálogo) → pega ruta → Enter → escribe etiqueta → aceptar → Enter
    for i, (ruta, etiqueta) in enumerate(paths):
        print(f"📎 Adjuntando: {etiqueta} → {ruta}")

        # Del 2do archivo en adelante: Enter abre el nuevo diálogo
        if i > 0:
            pyautogui.press("enter")
            pausa(2)

        win32clipboard.OpenClipboard()
        win32clipboard.EmptyClipboard()
        win32clipboard.SetClipboardText(ruta, win32clipboard.CF_UNICODETEXT)
        win32clipboard.CloseClipboard()
        pausa(1)
        pyautogui.hotkey("ctrl", "v")
        pausa(1)
        pyautogui.press("enter")
        pausa(1)

        pyautogui.write(etiqueta, interval=0.05)
        pausa(1)

        try:
            loc_aceptar = pyautogui.locateOnScreen(IMAGEN_ACEPTAR, confidence=CONFIDENCE)
        except Exception:
            loc_aceptar = None
        if not loc_aceptar:
            print(f"⚠ No se encontró aceptar.png para {etiqueta}")
        else:
            pyautogui.click(loc_aceptar)
            pausa(1)
            pyautogui.press("enter")
            pausa(1)

    print("✅ Adjuntos OLE procesados")
    return True

# ==========================================
# FLUJO CONFIGURACION
# ==========================================
def procesar_configuracion():
    doc_config = obtener_dato_de_columna("DOCUMENTACION")
    print(f"📋 DOC_CONFIG: {doc_config[:50] if doc_config else 'vacío'}...")

    # 0. Verificar paths y procesar adjuntos OLE si los hay
    paths = obtener_paths_de_excel()
    if paths:
        print(f"📂 Se encontraron {len(paths)} archivo(s) para adjuntar en OLE")
        procesar_adjuntos_ole(paths)
    else:
        print("📂 Todas las columnas PATH están en NO → saltando OLE")

    # 1. Buscar y click en tareas.png o tareas_1.png
    loc_tareas = None
    for img_tareas in [IMAGEN_TAREAS, IMAGEN_TAREAS_1]:
        try:
            loc_tareas = pyautogui.locateOnScreen(img_tareas, confidence=CONFIDENCE)
        except Exception:
            loc_tareas = None
        if loc_tareas:
            print(f"✅ Detectado: {img_tareas}")
            break

    if not loc_tareas:
        print("⚠ No se detectó tareas.png ni tareas_1.png")
        return False

    pyautogui.click(loc_tareas)
    pausa(1)
    pyautogui.press("enter")
    pausa(1)
    pyautogui.move(0, 30)
    pausa(1)

    # 2. Buscar primero, scrollear solo si no encuentra (3 scrolls de -5, 3 intentos por scroll)
    loc_config = None
    for scroll_num in range(3):
        for intento in range(3):
            try:
                loc_config = pyautogui.locateOnScreen(IMAGEN_CONFIGURACION, confidence=CONFIDENCE)
            except Exception:
                loc_config = None

            if loc_config:
                print("✅ Encontrado configuracion.png")
                break

            print(f"🔍 No se encontró configuracion.png (scroll {scroll_num + 1}/3, intento {intento + 1}/3), esperando...")
            pausa(2)

        if loc_config:
            break

        print("🔍 Haciendo scroll y reintentando...")
        pyautogui.scroll(-5)
        pausa(3)

    if not loc_config:
        print("⚠ No se encontró configuracion.png tras scrolls")
        return False

    # 3. Doble click directo con loc_config ya encontrado
    pyautogui.click(loc_config)
    pyautogui.click(loc_config) 
    pausa(3)

    # 4. Maximizar
    pyautogui.hotkey("win", "up")
    pausa(2)

    # Buscar editar_tarea.png
    try:
        loc_editar = pyautogui.locateOnScreen(IMAGEN_EDITAR_TAREA, confidence=CONFIDENCE)
    except Exception:
        loc_editar = None

    if not loc_editar:
        print("⚠ No se encontró editar_tarea.png")
        return False

    pyautogui.click(loc_editar)
    pausa(1)
    pyautogui.press("enter")
    pausa(1)

    # 5. Pegar documentación en anotaciones (ANTES de facturación)
    if not pegar_texto_en_anotaciones(doc_config, loc_config):
        return False

    # 6. Items de facturación (fechas con ESC + navegación)
    if not llenar_items_facturacion():
        return False

    return True

# ==========================================
# FLUJO PRINCIPAL
# ==========================================
def ejecutar_flujo(necesita_login):
    global flujo_en_ejecucion, last_activity_time

    with lock:
        if flujo_en_ejecucion:
            return
        flujo_en_ejecucion = True

    traer_crm_al_frente()
    print("⚡ Ejecutando flujo CONFIGURACION...")
    pausa(2)

    if necesita_login:
        user = os.environ.get("CRM_USER", "46243463").strip()
        pwd  = os.environ.get("CRM_PASS",  "AlaiaS02.*").strip()

        if not user or not pwd:
            print("❌ Faltan credenciales.")
            with lock:
                flujo_en_ejecucion = False
            return

        esperar_imagen_crm()
        pyautogui.press("tab"); pausa(1)
        pyautogui.press("tab"); pausa(1)
        pyautogui.write(user, interval=0.05); pausa(1)
        pyautogui.press("tab"); pausa(1)
        pyautogui.write(pwd,  interval=0.05); pausa(1)
        pyautogui.press("tab"); pausa(1)
        pyautogui.press("enter"); pausa(3)
        pyautogui.press("enter"); pausa(2)
        for _ in range(4):
            pyautogui.press("tab"); pausa(1)
        pyautogui.press("enter"); pausa(15)
        pyautogui.press("f2"); pausa(2)

    ok1 = esperar_imagen_con_f2()
    ok2 = copiar_pegar_otp()
    ok3 = detectar_flujo_y_maximizar()
    ok4 = procesar_configuracion()
    ok5 = guardar_crm()

    # Verificar OTP ocupada tras guardar
    try:
        loc_ocupada = pyautogui.locateOnScreen(IMAGEN_OTP_OCUPADA, confidence=CONFIDENCE)
    except Exception:
        loc_ocupada = None

    if loc_ocupada:
        print("⚠ OTP ocupada detectada después de guardar → presionando Enter")
        pyautogui.press("enter")
        pausa(1)
    else:
        print("✅ OTP libre luego de guardar")

    # Primer cierre
    try:
        loc_cierre = pyautogui.locateOnScreen(IMAGEN_CIERRE, confidence=CONFIDENCE)
        if loc_cierre:
            pyautogui.click(loc_cierre)
            pausa(1)
            print("✅ cierre.png clickeado")
    except Exception:
        print("⚠ No se encontró cierre.png")

    pyautogui.moveTo(1021, 517)
    pausa(0.5)

    # Segundo cierre
    try:
        loc_cierre = pyautogui.locateOnScreen(IMAGEN_CIERRE, confidence=CONFIDENCE)
        if loc_cierre:
            pyautogui.click(loc_cierre)
            pausa(1)
            print("✅ segundo cierre.png clickeado")
    except Exception:
        print("⚠ No se encontró segundo cierre.png")

    if ok1 and ok2 and ok3 and ok4 and ok5:
        marcar_fila_estado(fila_actual_excel, "COMPLETADO")
    else:
        marcar_fila_estado(fila_actual_excel, "ERROR")

    last_activity_time = time.time()
    with lock:
        flujo_en_ejecucion = False

    print("🏁 Flujo CONFIGURACION terminado")

# ==========================================
# INICIO
# ==========================================
if __name__ == "__main__":
    if not adquirir_lock():
        print("🟡 Ya hay un bot CONFIGURACION corriendo. Cerrando.")
    else:
        try:
            necesita_login_global = abrir_crm_si_no_existe()
            inicio_espera_sin_pendientes = None

            while True:
                time.sleep(1)

                if flujo_en_ejecucion:
                    inicio_espera_sin_pendientes = None
                    continue

                siguiente = obtener_siguiente_fila_pendiente()

                if siguiente:
                    inicio_espera_sin_pendientes = None
                    fila_actual_excel = siguiente
                    adquirir_cola_lock()
                    try:
                        ejecutar_flujo(necesita_login_global)
                    finally:
                        liberar_cola_lock()
                    necesita_login_global = False
                    continue

                if inicio_espera_sin_pendientes is None:
                    inicio_espera_sin_pendientes = time.time()
                    print("⏳ Sin pendientes CONFIGURACION. Esperando 8 minutos...")

                if time.time() - inicio_espera_sin_pendientes >= INACTIVITY_LIMIT:
                    print("🔴 8 minutos sin actividad. Cerrando CRM...")
                    try:
                        if crm_process:
                            crm_process.terminate()
                        else:
                            proc = buscar_crm()
                            if proc:
                                proc.terminate()
                        print("✅ CRM cerrado")
                    except Exception as e:
                        print(f"⚠ Error cerrando CRM: {e}")
                    break

        finally:
            liberar_lock()

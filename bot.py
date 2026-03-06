import subprocess
import psutil
import threading
import time
import pyautogui
import os
import openpyxl
import pyperclip
import win32clipboard

# ==========================================
# CONFIG
# ==========================================
CRM_PATH = r"C:\Users\dell\AppData\Roaming\Microsoft\Windows\Start Menu\Programs\Global Hitss\Global Hitss\Nuevo CRM.appref-ms"
IMAGEN_CONSULTAS = r"C:\Users\dell\Desktop\kickoff\project-Auto-bot\stratic\consultas.png"
IMAGEN_FLUJO_TRABAJO = r"C:\Users\dell\Desktop\kickoff\project-Auto-bot\stratic\flujo de trabajo.png"
IMAGEN_CRM = r"C:\Users\dell\Desktop\kickoff\project-Auto-bot\stratic\CRM.png"
IMAGEN_TAREAS = r"C:\Users\dell\Desktop\kickoff\project-Auto-bot\stratic\tareas.png"
IMAGEN_KICKOFF = r"C:\Users\dell\Desktop\kickoff\project-Auto-bot\stratic\Kickoff.png"
IMAGEN_KICKOFF_NOVEDADES = r"C:\Users\dell\Desktop\kickoff\project-Auto-bot\stratic\Kickoffnovedades.png"
IMAGEN_KICKOFF_NOVEDADES_GRIS = r"C:\Users\dell\Desktop\kickoff\project-Auto-bot\stratic\Kickoffnovedadesgris.png"
IMAGEN_OTH = r"C:\Users\dell\Desktop\kickoff\project-Auto-bot\stratic\oth_planear1.png"
IMAGEN_CIERRE = r"C:\Users\dell\Desktop\kickoff\project-Auto-bot\stratic\cierre.png"
IMAGEN_GRIS = r"C:\Users\dell\Desktop\kickoff\project-Auto-bot\stratic\gris.png"
IMAGEN_GUARDAR = r"C:\Users\dell\Desktop\kickoff\project-Auto-bot\stratic\guardar.png"
IMAGEN_ANOTACIONES = r"C:\Users\dell\Desktop\kickoff\project-Auto-bot\stratic\anotaciones.png"
IMAGEN_EDITAR_TAREA = r"C:\Users\dell\Desktop\kickoff\project-Auto-bot\stratic\editar_tarea.png"
IMAGEN_ASIGNACION_TAREA = r"C:\Users\dell\Desktop\kickoff\project-Auto-bot\stratic\asignacion_tarea.png"
IMAGEN_CODIGOS_RESOLUCION = r"C:\Users\dell\Desktop\kickoff\project-Auto-bot\stratic\codigos_de_resolucion.png"
IMAGEN_COD_RESOLUCION_1 = r"C:\Users\dell\Desktop\kickoff\project-Auto-bot\stratic\codigo_de_resolucion_1.png"
IMAGEN_FECHA_COMPROMISO = r"C:\Users\dell\Desktop\kickoff\project-Auto-bot\stratic\FECHA_compromiso.png"
IMAGEN_FECHA_PROGRAMACION = r"C:\Users\dell\Desktop\kickoff\project-Auto-bot\stratic\fecha_programacion1.png"
IMAGEN_TIPO_SERVICIO = r"C:\Users\dell\Desktop\kickoff\project-Auto-bot\stratic\tipo_de_servicio.png"
IMAGEN_ESTADO = r"C:\Users\dell\Desktop\kickoff\project-Auto-bot\stratic\estado.png"

EXCEL_PATH = r"C:\Users\dell\Desktop\kickoff\project-Auto-bot\datos.xlsx"
ESTADO_COLUMNA = "COMPLETADO"

INACTIVITY_LIMIT = 480
CONFIDENCE = 0.87

crm_process = None
last_activity_time = time.time()
flujo_en_ejecucion = False
lock = threading.Lock()

fila_actual_excel = None

pyautogui.FAILSAFE = True

# ==========================================
# LIMITE CRM (900 chars, Enter = 4)
# ==========================================
MAX_CARACTERES_CRM = 1000
ENTER_COST = 4

def dividir_texto_por_limite(texto, limite=MAX_CARACTERES_CRM):
    if texto is None:
        return []
    texto = str(texto).replace("\r\n", "\n").replace("\r", "\n")
    partes = []
    actual = []
    costo_actual = 0

    for linea in texto.split("\n"):
        c_linea = len(linea)
        c_enter = ENTER_COST if len(actual) > 0 else 0

        if (costo_actual + c_enter + c_linea) > limite and len(actual) > 0:
            partes.append("\n".join(actual))
            actual = []
            costo_actual = 0
            c_enter = 0

        if c_linea > limite:
            if len(actual) > 0:
                partes.append("\n".join(actual))
                actual = []
                costo_actual = 0
            start = 0
            while start < len(linea):
                end = min(start + limite, len(linea))
                partes.append(linea[start:end])
                start = end
            continue

        if c_enter:
            costo_actual += c_enter
        actual.append(linea)
        costo_actual += c_linea

    if len(actual) > 0:
        partes.append("\n".join(actual))

    return partes

# ====== LOCK PARA EVITAR 2 BOTS A LA VEZ ======
LOCK_FILE = os.path.join(os.path.dirname(__file__), "bot.lock")
lock_fd = None

def adquirir_lock():
    global lock_fd
    try:
        lock_fd = os.open(LOCK_FILE, os.O_CREAT | os.O_EXCL | os.O_RDWR)
        os.write(lock_fd, str(os.getpid()).encode())
        return True
    except FileExistsError:
        return False

def liberar_lock():
    global lock_fd
    try:
        if lock_fd is not None:
            os.close(lock_fd)
            lock_fd = None
        if os.path.exists(LOCK_FILE):
            os.remove(LOCK_FILE)
    except:
        pass

# ==========================================
# UTIL
# ==========================================
def pausa(seg=1):
    time.sleep(seg)

def obtener_marcacion_oth_de_excel():
    global fila_actual_excel
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb.active
        COL_D = 4  # D = MARCACION_OTH
        val = ws.cell(row=fila_actual_excel, column=COL_D).value
        if val is None:
            return ""
        return str(val).strip()
    except Exception as e:
        print(f"⚠ Error leyendo MARCACION_OTH (col D): {e}")
        return ""

def tabs_y_bajar_y_poner_cero():
    """
    TAB x6 -> DOWN -> escribir 0
    (se ejecuta DESPUES de hacer click en editar_tarea.png)
    """
    pyautogui.press("enter")
    pausa(1)

    marcacion = obtener_marcacion_oth_de_excel()

    pyautogui.press("tab")
    pausa(0.5)
    pyautogui.press("tab")

    pausa(0.2)
    if marcacion:
        win32clipboard.OpenClipboard()
        win32clipboard.EmptyClipboard()
        win32clipboard.SetClipboardText(marcacion, win32clipboard.CF_UNICODETEXT)
        win32clipboard.CloseClipboard()
        pyautogui.hotkey("ctrl", "v")
        pausa(0.2)
        print(f"✅ MARCACION_OTH escrita: {marcacion}")
    else:
        print("⚠ MARCACION_OTH vacía en Excel")

    for _ in range(4):
        pyautogui.press("tab")
        pausa(0.15)
    pyautogui.press("down")
    pausa(0.2)
    pyautogui.write("0", interval=0.05)
    pausa(0.2)
    pyautogui.press("tab")
    pausa(0.2)
    

def mostrar_escritorio():
    pyautogui.hotkey("win", "d")
    pausa(1)

# ==========================================
# BUSCAR CRM
# ==========================================
def buscar_crm():
    for proc in psutil.process_iter(['pid', 'name']):
        try:
            if proc.info['name'] and proc.info['name'].lower() == "crm.exe":
                return proc
        except:
            continue
    return None

# ==========================================
# TRAER CRM AL FRENTE
# ==========================================
def traer_crm_al_frente():
    ventanas = pyautogui.getAllWindows()
    for ventana in ventanas:
        if ventana.title and "CRM" in ventana.title.upper():
            ventana.restore()
            ventana.activate()
            pausa(1)
            print("🟢 CRM traído al frente")
            return True
    print("⚠ No se pudo traer CRM al frente")
    return False

# ==========================================
# ABRIR CRM SI NO EXISTE
# ==========================================
def abrir_crm_si_no_existe():
    global crm_process
    existente = buscar_crm()
    if existente:
        crm_process = existente
        traer_crm_al_frente()
        print("🟢 CRM ya estaba abierto")
        return False
    print("🚀 Abriendo CRM...")
    mostrar_escritorio()
    os.startfile(CRM_PATH)
    for _ in range(30):
        pausa(1)
        crm_process = buscar_crm()
        if crm_process:
            traer_crm_al_frente()
            print(f"✅ CRM abierto PID {crm_process.pid}")
            return True
    print("❌ No se pudo abrir CRM")
    return False

# ==========================================
# ESPERAR IMAGEN CON F2
# ==========================================
def esperar_imagen_con_f2():
    print("🔍 Buscando imagen consultas...")
    reintentos = 0
    max_reintentos = 5
    while True:
        traer_crm_al_frente()
        try:
            location = pyautogui.locateOnScreen(IMAGEN_CONSULTAS, confidence=CONFIDENCE)
        except Exception:
            location = None
        if location:
            print("✅ Imagen detectada")
            return True
        pyautogui.press("f2")
        pausa(2)
        reintentos += 1
        if reintentos >= max_reintentos:
            print("⚠ No se detectó la imagen después de 5 intentos, volviendo a intentar...")
            reintentos = 0

# ==========================================
# ESPERAR IMAGEN DEL CRM ANTES DE LOGIN
# ==========================================
def esperar_imagen_crm():
    print("🔍 Esperando que aparezca la imagen del CRM...")
    while True:
        try:
            location = pyautogui.locateOnScreen(IMAGEN_CRM, confidence=CONFIDENCE)
        except Exception:
            location = None
        if location:
            print("✅ Imagen del CRM detectada, continuando flujo")
            return True
        pausa(2)

# ==========================================
# EXCEL: ENCONTRAR SIGUIENTE FILA PENDIENTE
# ==========================================
def obtener_siguiente_fila_pendiente():
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb.active

        col_estado = None
        col_otp = None

        for idx, cell in enumerate(ws[1], 1):
            if cell.value == ESTADO_COLUMNA:
                col_estado = idx
            if cell.value == "OTP":
                col_otp = idx

        if not col_estado or not col_otp:
            return None

        for row in range(2, ws.max_row + 1):
            estado = ws.cell(row=row, column=col_estado).value
            otp_val = ws.cell(row=row, column=col_otp).value
            if (estado is None or str(estado).strip() == "") and (otp_val is not None and str(otp_val).strip() != ""):
                return row
        return None
    except:
        return None

# ==========================================
# EXCEL: OBTENER OTP DE UNA FILA
# ==========================================
def obtener_otp_de_fila(row):
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb.active

        col_otp = None
        for idx, cell in enumerate(ws[1], 1):
            if cell.value == "OTP":
                col_otp = idx
                break
        if not col_otp:
            return None

        val = ws.cell(row=row, column=col_otp).value
        if val is None:
            return None
        val = str(val).strip()
        if val == "":
            return None
        return val
    except:
        return None


def escribir_tipo_servicio_en_crm():
    # 1) 11 tabs
    for _ in range(11):
        pyautogui.press("tab")
        pausa(0.15)

    # 2) detectar imagen tipo_de_servicio.png
    try:
        loc = pyautogui.locateOnScreen(IMAGEN_TIPO_SERVICIO, confidence=CONFIDENCE)
    except Exception:
        loc = None

    if not loc:
        print("⚠ No se encontró tipo_de_servicio.png")
        return False

    x, y = pyautogui.center(loc)

    # 3) click +100px a la derecha
    pyautogui.click(x + 100, y)
    pausa(0.2)

    # 4) traer valor de Excel (col I)
    tipo = obtener_tipo_servicio_de_excel()
    if not tipo:
        print("⚠ TIPO_SERVICIO vacío en Excel")
        return True  # no es error crítico, solo no escribe

    # 5) escribir
    pyautogui.write(tipo, interval=0.05)
    pausa(0.2)

    print(f"✅ TIPO_SERVICIO escrito: {tipo}")
    return True

# ==========================================
# EXCEL: LEER COD RESOLUCION 1
# ==========================================
def leer_cod_resolucion_1_de_excel():
    try:
        import win32com.client
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        wb = excel.Workbooks.Open(EXCEL_PATH)
        ws = wb.ActiveSheet

        col_cod = None
        for col in range(1, 40):
            if ws.Cells(1, col).Value == "COD_RESOLUCION_1":
                col_cod = col
                break

        valor = None
        if col_cod:
            valor = ws.Cells(fila_actual_excel, col_cod).Value

        wb.Close(False)
        excel.Quit()

        if valor is None:
            return ""
        return str(valor).strip()
    except Exception as e:
        print(f"⚠ Error leyendo COD_RESOLUCION_1 desde Excel: {e}")
        return ""

# ==========================================
# EXCEL: LEER TEXTOS DE DOCUMENTACION
# ==========================================
def leer_textos_documentacion_de_excel():
    try:
        import win32com.client
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        wb = excel.Workbooks.Open(EXCEL_PATH)
        ws = wb.ActiveSheet

        texto_factibilidad = ws.Cells(fila_actual_excel, 2).Value
        texto_correo = ws.Cells(fila_actual_excel, 3).Value

        wb.Close(False)
        excel.Quit()
        return texto_factibilidad, texto_correo
    except Exception as e:
        print(f"⚠ Error leyendo textos desde Excel: {e}")
        return None, None

# ==========================================
# ✅ EXCEL (K/G/H) PARA FACTURACION
#   K = DOCUMENTACION_ITEM_FACTURACION
#   G = FECHA_COMPROMISO
#   H = FECHA_PROGRAMACION
# ==========================================
def obtener_datos_facturacion():
    global fila_actual_excel
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb.active

        COL_K = 11  # K
        COL_G = 7   # G
        COL_H = 8   # H

        flag = ws.cell(row=fila_actual_excel, column=COL_K).value
        if flag is None:
            return False, None, None

        flag = str(flag).strip().upper()
        if flag != "SI":
            return False, None, None

        fecha_compromiso = ws.cell(row=fila_actual_excel, column=COL_G).value
        fecha_programacion = ws.cell(row=fila_actual_excel, column=COL_H).value

        return True, fecha_compromiso, fecha_programacion
    except Exception as e:
        print(f"⚠ Error leyendo datos de FACTURACIÓN (K/G/H): {e}")
        return False, None, None

def obtener_tipo_servicio_de_excel():
    global fila_actual_excel
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb.active

        COL_I = 9  # I = TIPO_SERVICIO
        val = ws.cell(row=fila_actual_excel, column=COL_I).value

        if val is None:
            return ""
        return str(val).strip()
    except Exception as e:
        print(f"⚠ Error leyendo TIPO_SERVICIO (col I): {e}")
        return ""

def obtener_tipo_servicio_de_excel():
    global fila_actual_excel
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb.active

        COL_I = 9  # I = TIPO_SERVICIO
        val = ws.cell(row=fila_actual_excel, column=COL_I).value

        if val is None:
            return ""
        return str(val).strip()
    except Exception as e:
        print(f"⚠ Error leyendo TIPO_SERVICIO (col I): {e}")
        return ""

# ==========================================
# ✅ NUEVO: sacar dia/mes/año de FECHA_COMPROMISO
# ==========================================
def obtener_partes_fecha_compromiso(valor):
    try:
        from datetime import datetime, date

        if valor is None:
            return "", "", ""

        if isinstance(valor, datetime):
            d = valor.day
            m = valor.month
            y = valor.year
            return f"{d:02d}", f"{m:02d}", str(y)

        if isinstance(valor, date):
            d = valor.day
            m = valor.month
            y = valor.year
            return f"{d:02d}", f"{m:02d}", str(y)

        s = str(valor).strip()
        if not s:
            return "", "", ""

        # soporta "YYYY-MM-DD", "DD/MM/YYYY", "DD-MM-YYYY"
        if "-" in s:
            p = s.split("-")
            if len(p) == 3:
                if len(p[0]) == 4:   # YYYY-MM-DD
                    y, m, d = p[0], p[1], p[2]
                    return d.zfill(2), m.zfill(2), y
                else:                # DD-MM-YYYY
                    d, m, y = p[0], p[1], p[2]
                    return d.zfill(2), m.zfill(2), y
        if "/" in s:
            p = s.split("/")
            if len(p) == 3:          # DD/MM/YYYY
                d, m, y = p[0], p[1], p[2]
                return d.zfill(2), m.zfill(2), y

        return "", "", ""
    except:
        return "", "", ""
    

def obtener_gerencia_de_excel():
    global fila_actual_excel
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb.active

        COL_F = 6  # F = GERENCIA
        val = ws.cell(row=fila_actual_excel, column=COL_F).value

        if val is None:
            return ""

        return str(val).strip().upper()
    except Exception as e:
        print(f"⚠ Error leyendo GERENCIA (col F): {e}")
        return ""

# ==========================================
# ✅ MODIFICADO COMO PEDISTE:
# - localizar FECHA_compromiso.png
# - ESC
# - click centro +450px en X
# - escribir DIA, -> derecha, MES, -> derecha, AÑO (todo del compromiso)
# ==========================================
def click_fecha_compromiso_y_escribir_partes(fecha_compromiso, fecha_programacion):

    # ===============================
    # FECHA COMPROMISO
    # ===============================

    dia, mes, anio = obtener_partes_fecha_compromiso(fecha_compromiso)

    if not dia or not mes or not anio:
        print("⚠ FECHA_COMPROMISO inválida/vacía, no se escribe")
        return True

    try:
        loc = pyautogui.locateOnScreen(IMAGEN_FECHA_COMPROMISO, confidence=CONFIDENCE)
    except Exception:
        loc = None

    if not loc:
        print("⚠ No se encontró FECHA_compromiso.png")
        return False

    x, y = pyautogui.center(loc)

    pyautogui.click(x + 450, y)
    pausa(0.2)

    pyautogui.press("esc")
    pausa(0.2)

    pyautogui.write(dia, interval=0.05)
    pausa(0.1)
    pyautogui.press("right")
    pausa(0.1)

    pyautogui.write(mes, interval=0.05)
    pausa(0.1)
    pyautogui.press("right")
    pausa(0.1)

    pyautogui.write(anio, interval=0.05)
    pausa(0.2)

    pyautogui.press("tab")
    pausa(0.2)
    pyautogui.press("down")
    pausa(0.2)
    pyautogui.press("down")
    pausa(0.2)

    pyautogui.write("inteegra")
    pausa(0.2)

    pyautogui.press("tab")
    pausa(0.2)
    pyautogui.press("down")
    pausa(0.2)
    pyautogui.press("down")
    pausa(0.2)

    pyautogui.press("tab")
    pausa(0.2)
    pyautogui.press("down")
    pausa(0.2)
    pyautogui.press("down")
    pausa(0.2)

    pyautogui.write("kickoff")
    pausa(0.2)

    pyautogui.press("tab")
    pausa(0.2)
    pyautogui.press("down")
    pausa(0.2)
    pyautogui.press("down")
    pausa(0.2)

    ger = obtener_gerencia_de_excel()

    if "PROY" in ger:
        texto_kickoff = "kick off proyectos"
    elif "ESTAND" in ger:
        texto_kickoff = "kickoff estandar"
    else:
        texto_kickoff = "kickoff estandar"

    pyautogui.write(texto_kickoff)
    pausa(0.2)

    pyautogui.scroll(-30)
    pausa(0.2)

    pyautogui.press("tab")
    pausa(0.2)
    pyautogui.press("tab")
    pausa(0.2)
    pyautogui.press("tab")
    pausa(0.2)

    pyautogui.write("NO")
    pausa(0.5)

    print(f"✅ FECHA_COMPROMISO escrita: {dia}-{mes}-{anio}")

    
    pyautogui.press("tab")
    pausa(0.2)
    pyautogui.press("tab")
    pausa(0.2)
    pyautogui.press("tab")
    pausa(0.2)
    pyautogui.press("tab")
    pausa(1)

    # ===============================
    # FECHA PROGRAMACION
    # ===============================

    dia_p, mes_p, anio_p = obtener_partes_fecha_compromiso(fecha_programacion)

    if not dia_p or not mes_p or not anio_p:
        print("⚠ FECHA_PROGRAMACION inválida/vacía")
        return True

    try:
        loc_prog = pyautogui.locateOnScreen(IMAGEN_FECHA_PROGRAMACION, confidence=CONFIDENCE)
    except Exception:
        loc_prog = None

    if not loc_prog:
        print("⚠ No se encontró fecha_programacion1.png")
        return False
    
    pausa(1)

    xp, yp = pyautogui.center(loc_prog)

    pyautogui.click(xp + 450, yp)
    pausa(0.2)

    pausa(1)

    pyautogui.press("esc")
    pausa(0.2)

    pyautogui.write(dia_p, interval=0.05)
    pausa(0.1)
    pyautogui.press("right")
    pausa(0.1)

    pyautogui.write(mes_p, interval=0.05)
    pausa(0.1)
    pyautogui.press("right")
    pausa(0.1)

    pyautogui.write(anio_p, interval=0.05)
    pausa(0.2)

    print(f"✅ FECHA_PROGRAMACION escrita: {dia_p}-{mes_p}-{anio_p}")

    if not escribir_tipo_servicio_en_crm():
        return False

    return True


    

# ==========================================
# EXCEL: MARCAR FILA COMO COMPLETADO/ERROR
# ==========================================
def marcar_fila_estado(row, valor):
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb.active

        col_idx = None
        for idx, cell in enumerate(ws[1], 1):
            if cell.value == ESTADO_COLUMNA:
                col_idx = idx
                break

        if not col_idx:
            print(f"⚠ No se encontró columna {ESTADO_COLUMNA} en Excel")
            return

        ws.cell(row=row, column=col_idx).value = valor
        wb.save(EXCEL_PATH)
        print(f"✅ Excel fila {row} marcada como {valor}")
    except Exception as e:
        print(f"⚠ Error marcando Excel: {e}")

# ==========================================
# COPIAR OTP (DE LA FILA ACTUAL) Y PEGAR
# ==========================================
def copiar_pegar_otp():
    global fila_actual_excel
    print("📋 Copiando OTP desde Excel y pegando...")
    try:
        if not fila_actual_excel:
            print("⚠ No hay fila_actual_excel")
            return False

        otp = obtener_otp_de_fila(fila_actual_excel)
        if not otp:
            print("⚠ OTP vacío o no encontrado en la fila actual")
            return False

        pyperclip.copy(otp)
        pausa(0.3)
        pyautogui.hotkey("ctrl", "v")
        pausa(0.3)
        pyautogui.press("enter")

        print("✅ OTP pegado y Enter ejecutado")
        return True
    except Exception as e:
        print(f"⚠ Error en copiar_pegar_otp: {e}")
        return False

# ==========================================
# DETECTAR FLUJO Y MAXIMIZAR
# ==========================================
def detectar_flujo_y_maximizar():
    print("🔍 Buscando imagen flujo de trabajo...")
    reintentos = 0
    max_reintentos = 5
    while True:
        try:
            location = pyautogui.locateOnScreen(IMAGEN_FLUJO_TRABAJO, confidence=0.87)
        except Exception:
            location = None
        if location:
            print("✅ Flujo detectado → Maximizando ventana")
            pyautogui.hotkey("win", "up")
            pausa(2)
            return True
        pausa(2)
        reintentos += 1
        if reintentos >= max_reintentos:
            print("⚠ No se detectó el flujo después de 5 intentos, volviendo a intentar...")
            reintentos = 0

# ==========================================
# GUARDAR EN CRM
# ==========================================
def guardar():
    print("💾 Buscando botón guardar...")
    try:
        location_guardar = pyautogui.locateOnScreen(IMAGEN_GUARDAR, confidence=CONFIDENCE)
        if location_guardar:
            pyautogui.click(location_guardar)
            pausa(2)
            print("✅ Guardado correctamente")
            pyautogui.press("enter")
            pausa(1)
            pyautogui.press("enter")
            pausa(1)
            return True
        else:
            print("⚠ No se encontró guardar.png")
            return False
    except Exception as e:
        print(f"⚠ Error en guardar: {e}")
        return False

# ==========================================
# NOVEDADES: NORMAL O GRIS
# ==========================================
def buscar_kickoff_novedades(reintentos=5, pausa_entre=1.5):
    for intento in range(reintentos):
        print(f"🔍 Buscando Kickoffnovedades... intento {intento + 1}/{reintentos}")
        try:
            loc = pyautogui.locateOnScreen(IMAGEN_KICKOFF_NOVEDADES, confidence=CONFIDENCE)
            if loc:
                print("✅ Kickoffnovedades NORMAL detectado")
                return loc
            print(f"   ↳ NORMAL no encontrado, buscando GRIS...")
            loc = pyautogui.locateOnScreen(IMAGEN_KICKOFF_NOVEDADES_GRIS, confidence=CONFIDENCE)
            if loc:
                print("✅ Kickoffnovedades GRIS detectado")
                return loc
            print(f"   ↳ GRIS no encontrado")
        except Exception as e:
            print(f"   ↳ Excepción: {e}")
        pyautogui.scroll(-30)
        pausa(pausa_entre)

    print("❌ No se encontró Kickoffnovedades")
    return None

# ==========================================
# Buscar codigos_de_resolucion.png y luego COD_RESOLUCION_1 (TAB y escribir)
# ==========================================
def ejecutar_codigos_de_resolucion():
    try:
        loc_codigos = pyautogui.locateOnScreen(IMAGEN_CODIGOS_RESOLUCION, confidence=CONFIDENCE)
    except Exception:
        loc_codigos = None

    if not loc_codigos:
        print("⚠ No se encontró codigos_de_resolucion.png")
        return False

    pyautogui.click(loc_codigos)
    pausa(0.6)
    print("✅ codigos_de_resolucion.png clickeado")

    try:
        loc = pyautogui.locateOnScreen(IMAGEN_COD_RESOLUCION_1, confidence=CONFIDENCE)
    except Exception:
        loc = None

    if not loc:
        print("⚠ No se encontró codigo_de_resolucion_1.png")
        return False

    x, y = pyautogui.center(loc)
    pyautogui.click(x + 20, y)
    pausa(0.4)

    pyautogui.press("tab")
    pausa(0.35)

    valor_cod = leer_cod_resolucion_1_de_excel()

    if valor_cod:
        pyautogui.write(valor_cod, interval=0.05)
        print(f"✅ COD_RESOLUCION_1 escrito: {valor_cod}")
        return True

    print("⚠ COD_RESOLUCION_1 vacío en Excel")
    return True

# ==========================================
# PEGAR TEXTO EN ANOTACIONES (CON LIMITE + REINGRESO)
# ==========================================
def pegar_texto_en_campo(texto, location_item, reingreso_tipo):
    if texto is None or str(texto).strip() == "":
        return True

    x_item, y_item = pyautogui.center(location_item)
    print(f"📍 Coordenadas item original: x={x_item}, y={y_item}")

    partes = dividir_texto_por_limite(texto)

    for i, parte in enumerate(partes):
        try:
            location_anotaciones = pyautogui.locateOnScreen(IMAGEN_ANOTACIONES, confidence=CONFIDENCE)
        except Exception:
            location_anotaciones = None

        if not location_anotaciones:
            print("⚠ No se encontró anotaciones.png")
            return False

        x, y = pyautogui.center(location_anotaciones)
        pyautogui.moveTo(x + 20, y + 10)
        pyautogui.click()
        pausa(1)

        parte = str(parte).replace('\r\n', '\n').replace('\r', '\n')
        lineas = parte.split('\n')

        for j, linea in enumerate(lineas):
            if linea:
                win32clipboard.OpenClipboard()
                win32clipboard.EmptyClipboard()
                win32clipboard.SetClipboardText(linea, win32clipboard.CF_UNICODETEXT)
                win32clipboard.CloseClipboard()
                pyautogui.hotkey("ctrl", "v")
                pausa(0.1)

            if j < len(lineas) - 1:
                pyautogui.press("enter")
                pausa(0.1)

        pausa(0.3)
        pyautogui.press("enter")

        if i < len(partes) - 1:
            print("💾 Límite 900 alcanzado, guardando y reingresando...")

            if not guardar():
                return False

            try:
                loc_cierre = pyautogui.locateOnScreen(IMAGEN_CIERRE, confidence=CONFIDENCE)
            except Exception:
                loc_cierre = None
            if not loc_cierre:
                print("⚠ No se encontró cierre.png durante reintento por límite")
                return False
            pyautogui.click(loc_cierre)
            pausa(4)

            print(f"🔁 Reingresando con coordenadas originales: x={x_item}, y={y_item}")

            pyautogui.click(x_item, y_item)
            pyautogui.click(x_item, y_item); pausa(5)

            pyautogui.hotkey("win", "up")
            pausa(1)

            try:
                loc_editar = pyautogui.locateOnScreen(IMAGEN_EDITAR_TAREA, confidence=CONFIDENCE)
                if loc_editar:
                    pyautogui.click(loc_editar)
                    pausa(0.7)
                    tabs_y_bajar_y_poner_cero()
                    pausa(0.7)
                    ejecutar_codigos_de_resolucion()
            except:
                print("⚠ No se encontró editar_tarea.png en reingreso")

            pausa(2)

    return True

# ==========================================
# DESPUES de documentar -> si K == "SI":
# imprime fechas y llena FECHA_COMPROMISO por partes con ESC + click(x+450)
# ==========================================
def ejecutar_documentacion_para_item(location_item, tipo_item):
    texto_factibilidad, texto_correo = leer_textos_documentacion_de_excel()

    print("📝 Documentando FACTIBILIDAD (con límite 900)...")
    if not pegar_texto_en_campo(texto_factibilidad, location_item, tipo_item):
        return False

    print("📧 Documentando CORREO REPORTE INICIO (con límite 900)...")
    if not pegar_texto_en_campo(texto_correo, location_item, tipo_item):
        return False

    activar_fact, fecha_comp, fecha_prog = obtener_datos_facturacion()
    if activar_fact:
        print("🧾 DOCUMENTACION_ITEM_FACTURACION = SI")
        print(f"📅 FECHA_COMPROMISO (col G): {fecha_comp}")
        print(f"📅 FECHA_PROGRAMACION (col H): {fecha_prog}")
        if not click_fecha_compromiso_y_escribir_partes(fecha_comp,fecha_prog):
            return False
    else:
        print("⏭ DOCUMENTACION_ITEM_FACTURACION != SI → se salta FECHAS")

    # Validar CERRADO_OTP (col L) antes de escribir estado
    try:
        wb_c = openpyxl.load_workbook(EXCEL_PATH)
        ws_c = wb_c.active
        cerrado_otp = ws_c.cell(row=fila_actual_excel, column=12).value
        cerrado_otp = str(cerrado_otp).strip().upper() if cerrado_otp else ""
    except Exception as e:
        print(f"⚠ Error leyendo CERRADO_OTP (col L): {e}")
        cerrado_otp = ""

    # Detectar estado.png, click centro +200px, escribir CERRADO y Enter
    if cerrado_otp == "SI":
        try:
            loc_estado = pyautogui.locateOnScreen(IMAGEN_ESTADO, confidence=CONFIDENCE)
        except Exception:
            loc_estado = None

        if loc_estado:
            x_est, y_est = pyautogui.center(loc_estado)
            pyautogui.click(x_est + 200, y_est)
            pausa(0.3)
            pyautogui.write("CERRADO", interval=0.05)
            pausa(0.2)
            pyautogui.press("enter")
            pausa(0.3)
            print("✅ Estado CERRADO escrito")
        else:
            print("⚠ No se encontró estado.png")
    else:
        print("⏭ CERRADO_OTP != SI → se salta estado")

    return True

# ==========================================
# ABRIR ITEM + PREPARAR:
#   (DOBLE CLICK + WIN UP + EDITAR + TABS/DOWN/0 + CODIGOS_RESOLUCION + ASIGNACION)
# ==========================================
def abrir_item_y_preparar(location_item):
    if not location_item:
        return False

    pyautogui.scroll(-20)
    pausa(2)

    pyautogui.click(location_item)
    pyautogui.click(location_item)
    pausa(3)

    pyautogui.hotkey("win", "up")
    pausa(1)

    try:
        loc_editar = pyautogui.locateOnScreen(IMAGEN_EDITAR_TAREA, confidence=CONFIDENCE)
    except Exception:
        loc_editar = None

    if not loc_editar:
        print("⚠ No se encontró editar_tarea.png")
        return False

    pyautogui.click(loc_editar)
    pausa(1)

    tabs_y_bajar_y_poner_cero()
    pausa(0.6)

    ejecutar_codigos_de_resolucion()
    pausa(0.6)

    try:
        loc_asig = pyautogui.locateOnScreen(IMAGEN_ASIGNACION_TAREA, confidence=CONFIDENCE)
    except Exception:
        loc_asig = None
    if loc_asig:
        pyautogui.click(loc_asig)
        pyautogui.press("enter")
        pausa(5)
    else:
        print("⚠ No se encontró asignacion_tarea.png")

    return True

# ==========================================
# PROCESAR TAREAS (Kickoff / OTH / Novedades) + DOCUMENTAR
# ==========================================
def procesar_tareas_kickoff():
    try:
        try:
            location_tareas = pyautogui.locateOnScreen(IMAGEN_TAREAS, confidence=CONFIDENCE)
        except Exception:
            location_tareas = None

        if location_tareas:
            pyautogui.click(location_tareas)
            pausa(1)
            pyautogui.press("enter")
            pausa(1)
        else:
            print("⚠ No se detectó tareas.png, continuando flujo")

        pyautogui.move(0, 30)
        pausa(1)

        # 1) KICKOFF
        for _ in range(2):
            try:
                location_kickoff = pyautogui.locateOnScreen(IMAGEN_KICKOFF, confidence=CONFIDENCE)
            except Exception:
                location_kickoff = None

            if location_kickoff:
                print("✅ Se encontró Kickoff.png")
                if not abrir_item_y_preparar(location_kickoff):
                    return False
                if not ejecutar_documentacion_para_item(location_kickoff, "KICKOFF"):
                    return False
                return True
            pyautogui.scroll(-30)
            pausa(2)

        # 2) OTH
        print("❌ No se encontró Kickoff.png, buscando oth_planear1.png...")
        for _ in range(2):
            try:
                location_oth = pyautogui.locateOnScreen(IMAGEN_OTH, confidence=CONFIDENCE)
            except Exception:
                location_oth = None

            if location_oth:
                print("✅ Se encontró oth_planear1.png")
                if not abrir_item_y_preparar(location_oth):
                    return False
                if not ejecutar_documentacion_para_item(location_oth, "OTH"):
                    return False
                return True

            pyautogui.scroll(-30)
            pausa(2)

        # 3) NOVEDADES (NORMAL O GRIS)
        print("❌ Tampoco se encontró oth_planear1.png, buscando Kickoffnovedades (normal/gris)...")
        for _ in range(2):
            location_nov = buscar_kickoff_novedades()
            if location_nov:
                print("✅ Se encontró Kickoffnovedades (normal/gris)")
                if not abrir_item_y_preparar(location_nov):
                    return False
                if not ejecutar_documentacion_para_item(location_nov, "NOVEDADES"):
                    return False
                return True

            pyautogui.scroll(-30)
            pausa(2)

        print("❌ No se encontró Kickoff / OTH / Kickoffnovedades")
        try:
            location_cierre = pyautogui.locateOnScreen(IMAGEN_CIERRE, confidence=CONFIDENCE)
            if location_cierre:
                pyautogui.click(location_cierre)
                pausa(1)
        except:
            print("⚠ No se pudo hacer click en cierre.png")
        return False

    except Exception as e:
        print(f"⚠ Error en procesar_tareas_kickoff: {e}")
        return False

# ==========================================
# FLUJO PRINCIPAL (PROCESA 1 FILA)
# ==========================================
def ejecutar_flujo(necesita_login):
    global flujo_en_ejecucion, last_activity_time

    with lock:
        if flujo_en_ejecucion:
            return
        flujo_en_ejecucion = True

    traer_crm_al_frente()
    print("⚡ Ejecutando flujo...")
    pausa(2)

    if necesita_login:
        user = os.environ.get("CRM_USER", "46381705").strip()
        pwd = os.environ.get("CRM_PASS", "Haroa997**").strip()

        if not user or not pwd:
            print("❌ Faltan credenciales. Configura CRM_USER y CRM_PASS en variables de entorno.")
            with lock:
                flujo_en_ejecucion = False
            return

        esperar_imagen_crm()
        pyautogui.press("tab"); pausa(1)
        pyautogui.press("tab"); pausa(1)
        pyautogui.write(user, interval=0.05); pausa(1)
        pyautogui.press("tab"); pausa(1)
        pyautogui.write(pwd, interval=0.05); pausa(1)
        pyautogui.press("tab"); pausa(1)
        pyautogui.press("enter"); pausa(3)
        pyautogui.press("enter"); pausa(2)
        for _ in range(4):
            pyautogui.press("tab"); pausa(1)
        pyautogui.press("enter"); pausa(15)
        pyautogui.press("f2"); pausa(2)

    ok1 = esperar_imagen_con_f2()
    ok2 = copiar_pegar_otp()
    pausa(5)
    ok3 = detectar_flujo_y_maximizar()
    ok4 = procesar_tareas_kickoff()
    ok5 = guardar()

    try:
        loc_cierre = pyautogui.locateOnScreen(IMAGEN_CIERRE, confidence=CONFIDENCE)
        if loc_cierre:
            pyautogui.click(loc_cierre)
            pausa(1)
            print("✅ cierre.png detectado y clickeado")
    except Exception:
        print("⚠ No se encontró cierre.png")

    pyautogui.moveTo(1021, 517)
    pausa(0.5)

    try:
        loc_cierre = pyautogui.locateOnScreen(IMAGEN_CIERRE, confidence=CONFIDENCE)
        if loc_cierre:
            pyautogui.click(loc_cierre)
            pausa(1)
            print("✅ cierre.png detectado y clickeado")
    except Exception:
        print("⚠ No se encontró cierre.png")

    

    if ok1 and ok2 and ok3 and ok4 and ok5:
        marcar_fila_estado(fila_actual_excel, "COMPLETADO")
    else:
        marcar_fila_estado(fila_actual_excel, "ERROR")

    last_activity_time = time.time()
    with lock:
        flujo_en_ejecucion = False

    print("🏁 Flujo terminado correctamente")

# ==========================================
# INICIO (COLA POR EXCEL)
# ==========================================
if __name__ == "__main__":
    if not adquirir_lock():
        print("🟡 Ya hay un bot corriendo. Este se cierra para respetar la cola por Excel.")
    else:
        try:
            necesita_login_global = abrir_crm_si_no_existe()
            inicio_espera_sin_pendientes = None

            while True:
                time.sleep(1)

                if flujo_en_ejecucion:
                    inicio_espera_sin_pendientes = None
                    continue

                siguiente = obtener_siguiente_fila_pendiente()

                if siguiente:
                    inicio_espera_sin_pendientes = None
                    fila_actual_excel = siguiente
                    ejecutar_flujo(necesita_login_global)
                    necesita_login_global = False
                    continue

                if inicio_espera_sin_pendientes is None:
                    inicio_espera_sin_pendientes = time.time()
                    print("⏳ Sin pendientes. Iniciando espera de 8 minutos por nueva actividad...")

                if time.time() - inicio_espera_sin_pendientes >= INACTIVITY_LIMIT:
                    print("🔴 8 minutos sin actividad nueva. Cerrando CRM...")

                    try:
                        if crm_process:
                            crm_process.terminate()
                        else:
                            proc = buscar_crm()
                            if proc:
                                proc.terminate()
                        print("✅ CRM cerrado por inactividad")
                    except Exception as e:
                        print(f"⚠ Error cerrando CRM: {e}")

                    break

        finally:
            liberar_lock()